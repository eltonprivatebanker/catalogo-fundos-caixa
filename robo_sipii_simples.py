"""
ROBÔ SIPII CAIXA — v18.3 (Automação integral GitHub)
==========================================================
Novidades v18.3:
  + Corrige o vínculo de fundos com nomes abreviados no SIPII
  + Cria índices completos por código e CNPJ para metadados do fundos.json
  + Inclui fallback seguro para Hashdex Nasdaq Crypto e Carteira Quantitativa Quali
  + Mantém a busca aproximada anterior como compatibilidade

Novidades v18.2:
  + Sincroniza o gráfico IPCA 12M × meta com o acumulado oficial mais recente
  + Mantém uma única competência mensal na série usada pelo gráfico

Novidades v18.1:
  + Corrige duplicidade mensal do IPCA ao consolidar a série 433 por competência AAAA-MM
  + Remove o override incorreto de maio/2026 (0,50%); o valor oficial da série 433 é 0,58%
  + Limpa automaticamente duplicidades antigas da base local, priorizando dados oficiais

Novidades v18.0:
  + Atualiza automaticamente o fundos.json integral a cada execução
  + Exporta metadados operacionais/comerciais para dados_atuais.csv e Excel
  + Mantém fundos_caixa.json como índice leve de documentos por CNPJ
  + Corrige fallback de Lâmina (LA) separado do Boletim Comercial (LAC)
  + Inclui Termo de Adesão, benchmark, estratégia, captação, tributação, horários,
    adiantamento, mínimos, público-alvo, carência, ASG e observações

Histórico v17.1 vs v16:
  + Poupança nova e antiga com acum. ano pré-calculado server-side
    → cards.poupanca_nova: { valor, mensal, acum_ano, historico_ano, nota }
    → cards.poupanca_antiga: { valor, mensal, acum_ano, historico_ano, nota }
    → Série BCB 196 (nova); quando Selic>8,5% antiga=nova (mesma fórmula)
    → Série BCB 253 (TR mensal) só usada quando Selic≤8,5% (raro)
    → Dashboard exibe "Acum. ano" e texto descritivo da regra em desktop e mobile

Novidades v16 vs v15:
  + CDI acumulado 12M/24M/36M pré-calculado server-side (série 4391 BCB)
  + IPCA acumulado 24M/36M pré-calculado server-side (série 433 BCB)
  + PTAX histórico mensal (37M) pré-salvo em mercado_atual.json
  + fundos_caixa.json indexado por CNPJ → botões de documentos PDF no dashboard
  + Coluna 'codfundo' no CSV/Excel → links diretos Lâmina/Regulamento/etc.
  Mantém: CDI mensal real (série 4391), Focus OData+PDF+cache, SIPII scraping,
           kpis_dashboard.json, fallback SIPII, limpeza de backups.
"""

import json
import csv
import glob
import calendar
from pathlib import Path
from datetime import datetime, timedelta, date
import io
import time, unicodedata, traceback, re, requests
from bs4 import BeautifulSoup
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, InvalidSessionIdException, WebDriverException,
)

# ---------------------------------------------------------------------------
# Configurações Globais
# ---------------------------------------------------------------------------

URL_SIPII = "https://www.fundos.caixa.gov.br/sipii/pages/public/listar-fundos-internet.jsf"

PERFIS = [
    {"segmento": "PESSOA FÍSICA",   "sigla": "PF"},
    {"segmento": "PESSOA JURÍDICA", "sigla": "PJ"},
    {"segmento": "GOVERNO",         "sigla": "GOV"},
    {"segmento": "RPPS",            "sigla": "RPPS"},
    {"segmento": "TODOS",           "sigla": "TODOS"},
]

CATEGORIAS_FIXAS = [
    {"csv": "RENDA FIXA SIMPLES",            "texto_tela": "RENDA FIXA SIMPLES"},
    {"csv": "RENDA FIXA",                    "texto_tela": "RENDA FIXA"},
    {"csv": "RENDA FIXA REFERENCIADO",       "texto_tela": "RENDA FIXA REFERENCIADO"},
    {"csv": "RENDA FIXA CURTO PRAZO",        "texto_tela": "RENDA FIXA CURTO PRAZO"},
    {"csv": "MULTIMERCADO",                  "texto_tela": "MULTIMERCADO"},
    {"csv": "CAMBIAL",                       "texto_tela": "CAMBIAL"},
    {"csv": "ACOES",                         "texto_tela": "AÇÕES"},
    {"csv": "FUNDO DE INDICE",               "texto_tela": "FUNDO DE ÍNDICE"},
    {"csv": "FUNDOS MUTUOS DE PRIVATIZACAO", "texto_tela": "FUNDOS MÚTUOS DE PRIVATIZAÇÃO"},
]

CAIXA_LISTING_PAGES = [
    "https://www.caixa.gov.br/fundos-investimento/renda-fixa/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/referenciados/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/fundo-simples/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/curto-prazo/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/multimercado/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/cambiais/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/fundos-de-indices/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/voce/renda-fixa/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/empresa/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/fmp-fgts/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/rpps/Paginas/default.aspx",
]

MESES_PT = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]

DEBUG_COLUNAS = False

# ---------------------------------------------------------------------------
# Fechamentos confirmados / validação manual de virada de mês
# ---------------------------------------------------------------------------
# Estes valores são usados apenas para corrigir divergências conhecidas na virada
# de mês e evitar que valores parciais fiquem marcados como mês fechado.
# Chave no formato AAAA-MM.
FECHAMENTOS_CONFIRMADOS = {
    "cdi_mensal": {
        "2026-01": 1.16,
        "2026-02": 1.00,
        "2026-03": 1.21,
        "2026-04": 1.09,
        "2026-05": 1.07,
    },
    # IPCA deve vir da série oficial 433 do BCB.
    # Overrides só devem ser usados excepcionalmente e com valor oficialmente validado.
    "ipca_mensal": {},
    "ptax_mensal": {
        "2026-05": {"variacao_mes_fechado": 1.37},
    },
    "indices_mensais": {
        "^BVSP": {"2026-05": {"fechamento": 173787.0, "variacao_mes_fechado": -7.22}},
        "^GSPC": {"2026-05": {"variacao_mes_fechado": 5.15, "variacao_mes_fechado_brl": 6.59}},
        "^IXIC": {"2026-05": {"variacao_mes_fechado": 8.36}},
        "^DJI":  {"2026-05": {"variacao_mes_fechado": 1.99}},
        "^IFIX": {"2026-05": {"variacao_mes_fechado": None}},  # ★ v17.3: IFIX pendente
    },
}


# ---------------------------------------------------------------------------
# Caminhos de arquivo
# ---------------------------------------------------------------------------
BASE_DIR            = Path.cwd()
LOG_PATH            = BASE_DIR / "execucao.log"
IPCA_BASE_PATH      = BASE_DIR / "ipca_historico_base.json"
FOCUS_CACHE_PATH    = BASE_DIR / "focus_cache.json"
SELIC_BASE_PATH     = BASE_DIR / "historico da selic do BC.json"
META_INFLACAO_PATH  = BASE_DIR / "meta-vs-inflacao-efetiva.json"
FUNDOS_JSON_LOCAL   = BASE_DIR / "fundos.json"
FUNDOS_CAIXA_PATH   = BASE_DIR / "fundos_caixa.json"   # índice leve de documentos
FUNDOS_AUDITORIA_PATH = BASE_DIR / "auditoria_fundos.json"
PTAX_CACHE_PATH    = BASE_DIR / "ptax_historico_cache.json"

# ---------------------------------------------------------------------------
# Dicionário estático — URLs validadas manualmente
# ---------------------------------------------------------------------------
URL_ESTATICO = {
    "CAIXA BRASIL INFLACAO ATIVA FIF RF CRED PRIV": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/brasil-IPCA",
    "CAIXA CAPITAL PROTEGIDO IBOVESPA CICLICO I FIC FIF MM": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-capital-protegido-ibovespa-ciclico-1-multimercado",
    "CAIXA ETF IBOVESPA FUNDO DE INDICE": "https://www.caixa.gov.br/fundos-investimento/fundos-de-indices/etf-ibovespa-fundo-de-indice/",
    "CAIXA EXPERT CLARITAS VALOR FIC FIF ACOES -": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/caixa-expert-claritas-valor-fic-acoes/Paginas/default.aspx",
    "CAIXA EXPERT GIANT ZARATHUSTRA FIC MULTIMERCADO -": "https://www.caixa.gov.br/fundos-investimento/multimercado/expert-giant-zarathustra-fic-multimercado",
    "CAIXA FI BRASIL IMA-B TP RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-ima-b-titulos-rf-longo-prazo",
    "CAIXA FI BRASIL IRF-M 1 TP RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-irfm-1-titulos-publicos-rf",
    "CAIXA FI BRASIL IRF-M 1+ TP RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-irfm-1-titulos-publicos-rf-longo-prazo/",
    "CAIXA FIC ACOES EXPERT VERDE AM LONG BIAS -": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/CAIXA-Expert-Verde-Am-Long-Bias-FIC-Acoes/Paginas/default.aspx",
    "CAIXA FIC ACOES EXPERT VINCI VALOR DIVIDENDOS RPPS": "https://www.caixa.gov.br/fundos-investimento/rpps/caixa-fic-acoes-vinci-valor-dividendos-rpps",
    "CAIXA FIC ACOES EXPERT VINCI VALOR RPPS": "https://www.caixa.gov.br/fundos-investimento/rpps/caixa-fic-acoes-vinci-valor-rpps",
    "CAIXA FIC FIF ABSOLUTO PRE RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-absoluto-pre-rf-longo-prazo",
    "CAIXA FIC FIF ACOES IBOVESPA": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-ibovespa",
    "CAIXA FIC FIF ACOES MULTIGESTOR": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fic-fia-multigestor/Paginas/default.aspx",
    "CAIXA FIC FIF ALOCACAO MACRO MM LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-alocacao-macro-multimercado-longo-prazo/",
    "CAIXA FIC FIF BETA RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-beta-ref-di-lp",
    "CAIXA FIC FIF BRASIL DISPONIBILIDADES SIMPLES RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/brasil-disponibilidades",
    "CAIXA FIC FIF BRASIL GESTAO ESTRATEGICA RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/caixa-fic-brasil-gestao-estrategica-rf",
    "CAIXA FIC FIF BRASIL RF REFER DI LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/referenciados/fi-brasil-ref-di-longo-prazo",
    "CAIXA CAPITAL PROTEGIDO CESTA AGRO MM": "https://www.caixa.gov.br/fundos-investimento/multimercado/caixa_fic_fim_cap_protegido_cesta_agro/Paginas/default.aspx",
    "CAIXA FIC FIF CAPITAL IND PRECOS RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-capital-indice-de-precos-rf-longo-prazo",
    "CAIXA FIC FIF CLASSICO RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-classico-rf-longo-prazo",
    "CAIXA FIC FIF DESENVOLVER RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-desenvolver-rf-longo-prazo/",
    "CAIXA FIC FIF E-FUNDO RF LP": "https://www.caixa.gov.br/fundos-investimento/e-fundos/renda-fixa-longo-prazo",
    "CAIXA FIC FIF EMPREENDER RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-empreender-rf-longo-prazo/",
    "CAIXA FIC FIF EQUILIBRIO MPE RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/caixa-equilibrio-mpe-rf-lp/",
    "CAIXA FIC FIF ESTRATEGIA LIVRE MULTIMERCADO LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/estrategia-livre-mm-lp",
    "CAIXA FIC FIF ESTRATEGICO MULTIMERCADO LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-estrategico-multimercado-longo-prazo",
    "CAIXA FIC FIF EXECUTIVO RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-executivo-rf-longo-prazo",
    "CAIXA FIC FIF EXPERTISE RF CRED PRIV LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-expertise-rf-credito-privado-longo-prazo",
    "CAIXA FIC FIF FACIL RF SIMPLES": "https://www.caixa.gov.br/fundos-investimento/fundo-simples/fic-facil-rf-simples",
    "CAIXA FIC FIF FOCO IND PRECOS RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-foco-indice-precos-rf-longo-prazo",
    "CAIXA FIC FIF FOF SMART MULTIESTRATEGIA MM": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-multigestor",
    "CAIXA FIC FIF GERACAO JOVEM CRED PRIV RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-geracao-jovem-rf-credito-privado-longo-prazo",
    "CAIXA FIC FIF GIRO IMEDIATO RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-giro-imediato-ref-di-longo-prazo",
    "CAIXA FIC FIF HEDGE MULTIMERCADO LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/multimercado/caixa-fic-hedge-multimercado-lp/Paginas/default.aspx",
    "CAIXA FIC FIF IDEAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-ideal-rf-longo-prazo",
    "CAIXA FIC FIF INDEXA DOLAR CAMBIAL": "https://www.caixa.gov.br/fundos-investimento/cambiais/fic-cambial-dolar/Paginas/default.aspx",
    "CAIXA FIC FIF JUROS E MOEDAS MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-juros-moedas-multimercado-longo-prazo",
    "CAIXA FIC FIF JUROS E MOEDAS MM PLUS LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/caixa-fic-juros-e-moedas-multimercado-plus-lp/",
    "CAIXA FIC FIF NOVO BRASIL RF IMA-B LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-novo-brasil-ima-b-rf-longo-prazo",
    "CAIXA FIC FIF OAB RF CRED PRIV LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-oab-rf-credito-privado-longo-prazo",
    "CAIXA FIC FIF PATRIMONIO IND.DE PRECOS RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-patrimonio-indice-precos-rf-longo-prazo",
    "CAIXA FIC FIF PERSONAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-personal-rf-longo-prazo",
    "CAIXA FIC FIF PLENO REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-pleno-ref-di-longo-prazo",
    "CAIXA FIC FIF PRATICO RF CURTO PRAZO": "https://www.caixa.gov.br/fundos-investimento/curto-prazo/fic-pratico-curto-prazo",
    "CAIXA FIC FIF PREFERENCIAL REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-preferencial-ref-di-longo-prazo",
    "CAIXA FIC FIF RELACIONAMENTO PERSONAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-relacionamento-personal-rf-longo-prazo",
    "CAIXA FIC FIF RUBI RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-rubi-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF SELECAO RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-selecao-rf-longo-prazo",
    "CAIXA FIC FIF SIGMA RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-sigma-ref-di-longo-prazo",
    "CAIXA FIC FIF SOBERANO RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-soberano-rf-longo-prazo",
    "CAIXA FIC FIF SUPREMO RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-supremo-rf-longo-prazo",
    "CAIXA FIC FIF TITULO PUBLICO MPE RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/caixa-fic-tp-mpe-rf-lp/",
    "CAIXA FIC FIF TRANSFERENCIA VOLUNTARIA RF CP": "https://www.caixa.gov.br/fundos-investimento/curto-prazo/fic-transferencias-voluntarias-curto-prazo",
    "CAIXA FIC FIF TURQUESA CORPORATIVO RF CP": "https://www.caixa.gov.br/fundos-investimento/voce/renda-fixa/fic-turquesa-corporativo-curto-prazo",
    "CAIXA FIC FIFONLINE RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-investidor-rf-longo-prazo",
    "CAIXA FIC INVESTIDOR RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-investidor-rf-longo-prazo",
    "CAIXA FIC MULTIGESTOR GLOBAL EQUITIES IE": "https://www.caixa.gov.br/fundos-investimento/multimercado/caixa-multigestor-global-equities-invest-ext/Paginas/default.aspx",
    "CAIXA FIC RELACIONAMENTO IDEAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-relacionamento-ideal-rf-longo-prazo",
    "CAIXA FIF ACOES BDR NIVEL I": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-bdr-nivel-1",
    "CAIXA FIF ACOES BRASIL ETF IBOVESPA": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-brasil-etf-bovespa",
    "CAIXA FIF ACOES BRASIL IBOVESPA": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-brasil-ibovespa/Paginas/default.aspx",
    "CAIXA FIF ACOES BRASIL IBX-50": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-brasil-ibx-50",
    "CAIXA FIF ACOES CONSTRUCAO CIVIL": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-construcao-civil",
    "CAIXA FIF ACOES CONSUMO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-consumo",
    "CAIXA FIF ACOES DIVIDENDOS": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-dividendos",
    "CAIXA FIF ACOES IBOVESPA ATIVO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-ibovespa-ativo",
    "CAIXA FIF ACOES IBRX ATIVO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-ibrx-ativo",
    "CAIXA FIF ACOES INDEXA PIBB IBRX 50": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-pibb-sem-opcao-de-venda",
    "CAIXA FIF ACOES INDEXA SETOR FINANCEIRO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/caixa-fia-indexa-setor-financeiro/Paginas/default.aspx",
    "CAIXA FIF ACOES INFRAESTRUTURA": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-infraestrutura",
    "CAIXA FIF ACOES INSTITUCIONAL BDR NIVEL I": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/FIA-Institucional-BDR-nivel-I/Paginas/default.aspx",
    "CAIXA FIF ACOES PETROBRAS": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-petrobras",
    "CAIXA FIF ACOES PETROBRAS PLUS": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fia-caixa-petrobras-plus/Paginas/default.aspx",
    "CAIXA FIF ACOES PETROBRAS PRE-SAL": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-petrobras-pre-sal",
    "CAIXA FIF ACOES SMALL CAPS ATIVO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-small-caps-ativo",
    "CAIXA FIF ACOES SUSTENTABILIDADE EMPRESARIAL ISE - IS": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-ise",
    "CAIXA FIF ACOES VALE DO RIO DOCE": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-vale-rio-doce",
    "CAIXA FIF ALIANCA TP RF CURTO PRAZO -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-alianca-rf",
    "CAIXA FIF BRASIL IDKA IPCA 2A TIT PUB RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-idka-ipca-2a-rf-longo-prazo",
    "CAIXA FIF BRASIL IMA GERAL TP RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-ima-geral-tp-rf-longo-prazo",
    "CAIXA FIF BRASIL IMA-B 5 TP RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-ima-b-5-titulos-publicos-rf-longo-prazo",
    "CAIXA FIF BRASIL IMA-B 5+ TP RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-ima-b-5_mais-titulos-publicos-rf-lp/Paginas/default.aspx",
    "CAIXA FIF BRASIL IRF-M TP RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/brasil-irf-m-titulos-publicos-renda-fixa-longo-prazo",
    "CAIXA FIF BRASIL TITULOS PUBLICOS RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-titulos-publicos-rf-longo-prazo",
    "CAIXA FIF DIAMANTE CORP RF CRED PRIV LP -": "https://www.caixa.gov.br/fundos-investimento/empresa/renda-fixa/credito-privado/caixa-fi-diamante-corporativo-rf-cred-priv-lp",
    "CAIXA FIF E-SIMPLES RENDA FIXA LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-renda-fixa-simples-lp/",
    "CAIXA FIF INDEXA BOLSA AMERICANA MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-bolsa-americana-multimercado-lp",
    "CAIXA FIF INDEXA OURO MULTIMERCADO LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-ouro-multimercado-longo-prazo",
    "CAIXA FIF MULTIMERCADO RV 30 LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-multimercado-rv30-longo-prazo",
    "CAIXA FIF RS TITULOS PUBLICOS RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-rs-titulos-publicos-rf-longo-prazo",
    "CAIXA FIF SAUDE SUPLEMENTAR ANS II RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-saude-suplementar-ans-ii-rf-longo-prazo",
    "CAIXA FIF SAUDE SUPLEMENTAR ANS RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-saude-suplementar-anf-rf",
    "FIC FIF CAIXA EXPERT BTG PACTUAL X10 MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/btg-pactual-x10-multimercado-longo-prazo/",
    "CAIXA FIC FIF OBJETIVO PRE RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-objetivo-pre-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF PERFORMANCE IMA-B RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-performance-ima-b-renda-fixa-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF PLUS QUALI RF CREDI PRIV LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-plus-qualificado-rf-credito-privado-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF ESPECIAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-especial-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF FIDELIDADE PRIVATE RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-fidelidade-private-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF MAXI RENDA FIXA CRED PRIV LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-maxi-rf-credito-privado-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF FIDELIDADE II RF CRED PRIV LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-fidelidade-ii-rf-credito-privado-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC OMEGA REF DI RF LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-omega-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF MEGA RF REFERENC DI LONGO PRAZO -": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-mega-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF TOP PRIVATE RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-top-private-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF FOF SMART CRED PRIV MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-fof-smart-credito-privado-multimercado-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF INDEXA EURO MM LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-euro-multimercado-longo-prazo/Paginas/default.aspx",
    "CAIXA CAPITAL PROTEGIDO CICLICO III FIC FIF MM LP -": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-capital-protegido-ciclico-iii-multimercado-lp/Paginas/default.aspx",
    "CAIXA FIC FIM CAPITAL PROTEGIDO CICLICO II LP - RL": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-capital-protegido-ciclico-ii-multimercado-lp/Paginas/default.aspx",
    "CAIXA FIC FIF EXPERT PIMCO INCOME IE MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/caixa-expert-pimco-income-ie-fic-multimercado/Paginas/default.aspx",
    "CAIXA FIF EXTRAMERCADO COMUM IRFM-1 RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-extramercado-comum-irfm-1-rf/Paginas/default.aspx",
    "CAIXA FIC FIF AMETISTA CORP RF SIMPLES": "https://www.caixa.gov.br/fundos-investimento/fundo-simples/fic-ametista-corporativo-rf-simples/Paginas/default.aspx",
    "CAIXA GIRO EMPRESAS FIC FIF RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-giro-empresas-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF CNI RF LP -": "https://www.caixa.gov.br/fundos-investimento/empresa/renda-fixa/fi-cni-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF SEBRAE RF LP": "https://www.caixa.gov.br/fundos-investimento/empresa/renda-fixa/fi-sebrae-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF GIRO MPE REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-giro-mpe-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF TOPAZIO CORP RF REFERENC DI -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/referenciados/fic-topazio-rf-ref-di-lp/Paginas/default.aspx",
    "CAIXA FIC ESMERALDA CORP RF REF DI CRED PRIV LP -": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-esmeralda-corp-ref-di/Paginas/default.aspx",
}

# ---------------------------------------------------------------------------
# Utilidades gerais
# ---------------------------------------------------------------------------
def log(msg):
    linha = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    print(linha)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(linha + "\n")

def normalizar(txt):
    txt = txt or ""
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return " ".join(txt.upper().split())

def chave_estatica(nome):
    n = normalizar(nome)
    n = re.sub(r'\s*RESP\s*LTDA.*$', '', n)
    n = re.sub(r'\s*-\s*RL$', '', n)
    n = re.sub(r'\s*\(\d+\).*$', '', n)
    return n.strip()

def formatar_cnpj(cnpj_raw):
    if not cnpj_raw:
        return ""
    digits = re.sub(r'\D', '', str(cnpj_raw))
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    return str(cnpj_raw)

TEXTO_PARA_CSV = {normalizar(c["texto_tela"]): c["csv"] for c in CATEGORIAS_FIXAS}

# ---------------------------------------------------------------------------
# Camada 2 — Scraping dinâmico de URLs da CAIXA
# ---------------------------------------------------------------------------
STOPWORDS = {
    "CAIXA","FIC","FIF","FI","RF","LP","MM","IE","RL","IS",
    "RESP","LTDA","DE","DO","DA","DOS","DAS","E","A","O","EM",
    "FUNDO","FUNDOS","RENDA","FIXA","LONGO","PRAZO","CREDITO",
    "PRIVADO","TITULOS","PUBLICOS","TP","REFERENCIADO","SIMPLES",
}

def palavras_chave(texto):
    n = normalizar(texto)
    n = re.sub(r'\s*RESP\s*LTDA.*$', '', n)
    n = re.sub(r'\s*\(\d+\).*$', '', n)
    n = n.replace('-', ' ')
    tokens = set(n.split())
    return tokens - STOPWORDS - {t for t in tokens if len(t) <= 2}

def palavras_slug(url):
    try:
        path = url.rstrip('/').split('?')[0]
        parte = path.split('/')[-1]
        if parte.lower() in ('default.aspx', ''):
            parte = path.split('/')[-2]
        parte = re.sub(r'\.aspx$', '', parte).replace('-', ' ').replace('_', ' ')
        tokens = set(normalizar(parte).split())
        return tokens - STOPWORDS - {t for t in tokens if len(t) <= 2}
    except:
        return set()

HEADERS_HTTP = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/123.0.0.0 Safari/537.36",
    "Accept-Language": "pt-BR,pt;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.caixa.gov.br/",
}

def raspar_urls_caixa():
    registros, vistos = [], set()
    base = "https://www.caixa.gov.br"
    session = requests.Session()
    session.headers.update(HEADERS_HTTP)
    for page_url in CAIXA_LISTING_PAGES:
        try:
            resp = session.get(page_url, timeout=20)
            if resp.status_code != 200:
                log(f"  [CAIXA] {page_url.split('/')[-3]} → HTTP {resp.status_code}")
                continue
            soup = BeautifulSoup(resp.text, "html.parser")
            novos = 0
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                texto = a.get_text(separator=" ", strip=True)
                if href.startswith("/"): href = base + href
                if not href.startswith("http"): continue
                if "/fundos-investimento/" not in href: continue
                if len(href.rstrip('/').split('/')) < 7: continue
                if href in vistos: continue
                vistos.add(href)
                registros.append({
                    "texto": normalizar(texto),
                    "url": href,
                    "palavras_texto": palavras_chave(texto),
                    "palavras_slug":  palavras_slug(href),
                })
                novos += 1
            log(f"  [CAIXA] {page_url.split('/')[-3]} → +{novos} links ({len(registros)} total)")
            time.sleep(1)
        except Exception as e:
            log(f"  [CAIXA] Erro ao raspar {page_url}: {e}")
    log(f"[CAIXA] Total URLs dinâmicas: {len(registros)}")
    return registros

def encontrar_url_dinamica(nome, registros):
    pf = palavras_chave(nome)
    if not pf: return ""
    melhor_url, melhor_score = "", 0.0
    for reg in registros:
        s = max(
            len(pf & reg["palavras_texto"]) / len(pf),
            len(pf & reg["palavras_slug"])  / len(pf),
        )
        if s > melhor_score:
            melhor_score, melhor_url = s, reg["url"]
    return melhor_url if melhor_score >= 0.70 else ""

def encontrar_url(nome, registros_dinamicos):
    chave = chave_estatica(nome)
    if chave in URL_ESTATICO:
        return URL_ESTATICO[chave]
    chave_curta = re.sub(r'\s+-\s*$', '', chave).strip()
    if chave_curta in URL_ESTATICO:
        return URL_ESTATICO[chave_curta]
    return encontrar_url_dinamica(nome, registros_dinamicos)

# ---------------------------------------------------------------------------
# Boletim Focus (BCB OData + fallback PDF + cache local) — inalterado v15
# ---------------------------------------------------------------------------
FOCUS_BASE = "https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata/ExpectativasMercadoAnuais"
INDICADORES_FOCUS = ["IPCA", "Selic", "PIB Total", "Câmbio", "IGP-M"]

def _buscar_focus_indicador(indicador, anos, headers):
    if indicador in ("IPCA", "IPCA Modal"):
        filtro = f"Indicador eq '{indicador}' and baseCalculo eq 0"
    else:
        filtro = f"Indicador eq '{indicador}'"
    url = (f"{FOCUS_BASE}?$filter={requests.utils.quote(filtro)}"
           f"&$format=json&$orderby=Data%20desc&$top=80")
    resultado = {}
    try:
        res = requests.get(url, headers=headers, timeout=15)
        if res.status_code != 200:
            log(f"  [Focus] {indicador} → HTTP {res.status_code}")
            return resultado
        registros = res.json().get("value", [])
        vistos_anos = set()
        for reg in registros:
            ano_str = str(reg.get("DataReferencia", ""))[:4]
            if not ano_str: continue
            try: ano = int(ano_str)
            except ValueError: continue
            if ano not in anos or ano in vistos_anos: continue
            vistos_anos.add(ano)
            resultado[ano] = {
                "mediana": reg.get("Mediana"),
                "media":   reg.get("Media"),
                "minimo":  reg.get("Minimo"),
                "maximo":  reg.get("Maximo"),
                "data_ref": reg.get("Data"),
            }
            if len(vistos_anos) == len(anos): break
    except Exception as e:
        log(f"  [Focus] Erro ao buscar {indicador}: {e}")
    return resultado

_FOCUS_PDF_INDICES = {2026: 2, 2027: 6, 2028: 10, 2029: 13}
_FOCUS_PDF_PADROES = {
    "IPCA":      r"IPCA\s*\(varia[çc][aã]o\s*%\)",
    "Selic":     r"Selic\s*\(%\s*a\.a[\).]?",
    "PIB Total": r"PIB\s*Total",
    "Câmbio":    r"C[âa]mbio\s*\(R\$\/US\$\)",
    "IGP-M":     r"IGP-M\s*\(varia[çc][aã]o\s*%\)",
}

def _ultima_sexta(offset_semanas=0):
    hoje = datetime.now()
    dias = (hoje.weekday() - 4) % 7
    data = hoje - timedelta(days=dias + offset_semanas * 7)
    return data.strftime("%Y%m%d"), data.strftime("%d/%m/%Y")

def _extrair_decimais_linha(texto):
    nums = re.findall(r'\b\d+[,\.]\d+\b', texto)
    resultado = []
    for n in nums:
        try:
            val = float(n.replace(',', '.'))
            if 0 < val < 500:
                resultado.append(val)
        except Exception:
            pass
    return resultado

def _parsear_texto_focus(texto, data_str):
    anos = [2026, 2027, 2028, 2029]
    resultado = {}
    data_ref = f"{data_str[:4]}-{data_str[4:6]}-{data_str[6:]}"
    for indicador, padrao in _FOCUS_PDF_PADROES.items():
        match = re.search(padrao, texto, re.IGNORECASE)
        if not match: continue
        trecho = texto[match.end():match.end() + 700]
        decimais = _extrair_decimais_linha(trecho)
        if len(decimais) < 11: continue
        resultado[indicador] = {}
        for ano in anos:
            idx = _FOCUS_PDF_INDICES[ano]
            if idx < len(decimais):
                val = round(decimais[idx], 4)
                resultado[indicador][ano] = {
                    "mediana": val, "media": val,
                    "minimo": None, "maximo": None,
                    "data_ref": data_ref,
                }
    return resultado if resultado else None

def _descobrir_focus_pdf_disponivel(headers, max_semanas=8):
    """
    Localiza o PDF do Boletim Focus mais recente realmente disponível no site do BCB.

    Motivo:
    Nem sempre o PDF da última sexta-feira já está publicado quando o robô ou a página
    rodam. Por isso não é seguro montar o link apenas com a data do computador.
    """
    log("[Focus PDF] Procurando PDF oficial mais recente disponível...")

    for offset in range(max_semanas):
        data_str, data_br = _ultima_sexta(offset)
        url_pdf = f"https://www.bcb.gov.br/content/focus/focus/R{data_str}.pdf"

        try:
            res = requests.get(url_pdf, headers=headers, timeout=20)
            content_type = res.headers.get("Content-Type", "").lower()
            tamanho = len(res.content or b"")

            if res.status_code == 200 and ("pdf" in content_type or tamanho > 10000):
                log(f"[Focus PDF] PDF disponível encontrado: R{data_str}.pdf ({data_br})")
                return {
                    "data_pdf": data_str,
                    "data_pdf_br": data_br,
                    "pdf_url": url_pdf,
                }

            log(f"[Focus PDF] R{data_str}.pdf indisponível — HTTP {res.status_code} | {tamanho} bytes")

        except Exception as e:
            log(f"[Focus PDF] Erro ao testar R{data_str}.pdf: {e}")

    log("[Focus PDF] Nenhum PDF disponível encontrado nas últimas semanas.")
    return {}


def _raspar_focus_pdf(headers):
    try:
        import pypdf
    except ImportError:
        log("[Focus PDF] pypdf não instalado — pip install pypdf")
        return None, None
    for offset in range(3):
        data_str, data_br = _ultima_sexta(offset)
        url_pdf = f"https://www.bcb.gov.br/content/focus/focus/R{data_str}.pdf"
        try:
            log(f"[Focus PDF] Tentando R{data_str}.pdf ({data_br})...")
            res = requests.get(url_pdf, headers=headers, timeout=35)
            if res.status_code != 200 or len(res.content) < 10000: continue
            reader = pypdf.PdfReader(io.BytesIO(res.content))
            texto = "".join((p.extract_text() or "") + "\n" for p in reader.pages)
            if "IPCA" not in texto or "Selic" not in texto: continue
            dados = _parsear_texto_focus(texto, data_str)
            if dados: return dados, data_str
        except Exception as e:
            log(f"[Focus PDF] Erro com R{data_str}.pdf: {e}")
    return None, None

def _salvar_focus_cache(focus):
    try:
        FOCUS_CACHE_PATH.write_text(json.dumps(focus, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        log(f"[Focus] Erro ao salvar cache: {e}")

def _carregar_focus_cache():
    if FOCUS_CACHE_PATH.exists():
        try:
            dados = json.loads(FOCUS_CACHE_PATH.read_text(encoding="utf-8"))
            log(f"[Focus] Cache local carregado — coletado em: {dados.get('data_coleta','?')}")
            return dados
        except Exception as e:
            log(f"[Focus] Erro ao ler cache: {e}")
    return None

def buscar_focus(headers):
    anos_alvo = [2026, 2027, 2028, 2029]
    log("[Focus] Coletando expectativas do Banco Central...")

    # Localiza o PDF oficial mais recente disponível no BCB.
    # Esse campo será usado pelo index.html para o botão "Baixe aqui o PDF",
    # evitando links para arquivos ainda não publicados.
    pdf_disponivel = _descobrir_focus_pdf_disponivel(headers)

    raw_focus = {}
    falhas_api = 0

    for indicador in INDICADORES_FOCUS:
        log(f"  [Focus] -> {indicador}")
        obtido = False

        for tentativa in range(2):
            resultado = _buscar_focus_indicador(indicador, anos_alvo, headers)

            if resultado:
                raw_focus[indicador] = resultado
                obtido = True
                falhas_api = 0
                break

            if tentativa < 1:
                time.sleep(10)

        if not obtido:
            raw_focus[indicador] = {}
            falhas_api += 1

            if falhas_api >= 2:
                log("[Focus] BCB indisponível — abortando OData.")
                break

        time.sleep(1)

    indicadores_ok = sum(1 for v in raw_focus.values() if v)

    # Fallback: se a API OData não trouxer nada, tenta extrair do PDF.
    if indicadores_ok == 0:
        dados_pdf, data_pdf = _raspar_focus_pdf(headers)

        if dados_pdf:
            url_pdf = f"https://www.bcb.gov.br/content/focus/focus/R{data_pdf}.pdf"
            data_pdf_br = f"{data_pdf[6:]}/{data_pdf[4:6]}/{data_pdf[:4]}"

            focus_pdf = {
                "data_coleta": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "fonte": "pdf",
                "data_pdf": data_pdf,
                "data_pdf_br": data_pdf_br,
                "pdf_url": url_pdf,
                "IPCA": dados_pdf.get("IPCA", {}),
                "Selic": dados_pdf.get("Selic", {}),
                "PIB": dados_pdf.get("PIB Total", {}),
                "Cambio": dados_pdf.get("Câmbio", {}),
                "IGPM": dados_pdf.get("IGP-M", {}),
                "IPCA_Modal": dados_pdf.get("IPCA", {}),
            }

            _salvar_focus_cache(focus_pdf)
            return focus_pdf

    if indicadores_ok == 0:
        cache = _carregar_focus_cache()

        if cache:
            cache["fonte"] = "cache"

            # Mesmo quando usa cache, tenta atualizar apenas o link do PDF mais recente disponível.
            if pdf_disponivel:
                cache.update(pdf_disponivel)

            return cache

    focus = {
        "data_coleta": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "fonte": "odata",
        "IPCA":       raw_focus.get("IPCA", {}),
        "Selic":      raw_focus.get("Selic", {}),
        "PIB":        raw_focus.get("PIB Total", {}),
        "Cambio":     raw_focus.get("Câmbio", {}),
        "IGPM":       raw_focus.get("IGP-M", {}),
        "IPCA_Modal": raw_focus.get("IPCA", {}),
    }

    # Acrescenta link validado do PDF oficial mais recente.
    if pdf_disponivel:
        focus.update(pdf_disponivel)

    if indicadores_ok > 0:
        _salvar_focus_cache(focus)

    return focus


# ---------------------------------------------------------------------------
# ★ fundos.json — metadados comerciais + fundos_caixa.json para o HTML (v16)
# ---------------------------------------------------------------------------
FUNDOS_JSON_URL = "https://www.caixa.gov.br/CAIXA-Asset/Documents/data/fundos.json"
INDISPONIVEL = {"INDISPONIVEL", "indisponivel", "", None}

_STOPWORDS_MATCH = {
    "CAIXA","FIC","FIF","FI","RF","LP","MM","IE","RL","IS","TP",
    "RESP","LTDA","DE","DO","DA","DOS","DAS","E","A","O","EM",
    "FUNDO","FUNDOS","RENDA","FIXA","LONGO","PRAZO","CREDITO",
    "PRIVADO","TITULOS","PUBLICOS","REFERENCIADO","SIMPLES","CURTO",
}

def _normalizar_nome_fundo(nome):
    n = unicodedata.normalize("NFD", str(nome).upper())
    n = "".join(c for c in n if unicodedata.category(c) != "Mn")
    n = re.sub(r"[^A-Z0-9 ]", " ", n)
    return re.sub(r"\s+", " ", n).strip()

def _palavras_chave_fundo(nome):
    n = _normalizar_nome_fundo(nome)
    n = re.sub(r"\bRESP LTDA\b", "", n)
    n = re.sub(r"\bRL\b", "", n)
    tokens = set(re.sub(r"\s+", " ", n).strip().split())
    return tokens - _STOPWORDS_MATCH - {t for t in tokens if len(t) <= 2}

def _url_valida(url):
    if not url or str(url).strip().upper() in INDISPONIVEL:
        return False
    url = str(url).strip()
    return url.startswith("http") and "caixa.gov.br" in url and "/Sistemas/" not in url

def _extrair_codfundo(item):
    """
    ★ v16.2 — usa co_siico00 (campo real confirmado pelo log de debug).
    Fallback: extrai código dos URLs de documentos presentes no JSON.
    """
    # Campo principal confirmado: co_siico00='0054', co_siico='54'
    for campo in ["co_siico00", "co_siico", "nu_fundo", "co_fundo", "cd_fundo"]:
        val = item.get(campo)
        if val is not None:
            s = re.sub(r"[^0-9]", "", str(val).strip())
            if 2 <= len(s) <= 6:
                return s.lstrip("0") or s

    # Fallback: extrai código dos URLs de documentos (ex: RG_7880.pdf → 7880)
    for campo_url in ["de_link_regulamento", "de_link_lamina", "de_link_info_compl",
                      "de_link_fato_relevante", "de_link_pagina_fundo"]:
        url = str(item.get(campo_url) or "").strip()
        if not url or "INDISPONIVEL" in url.upper():
            continue
        m = re.search(r"[/_]([A-Z]+)_([0-9]{2,6})[._]", url)
        if m:
            return m.group(2)
    return ""

def _url_doc_valida(url):
    """Verifica se uma URL de documento é usável."""
    if not url or not str(url).strip():
        return False
    u = str(url).strip().upper()
    return u not in ("INDISPONIVEL", "NULL", "NONE", "") and str(url).startswith("http")


def _salvar_json_atomico(caminho, dados):
    """Grava JSON sem risco de deixar arquivo parcial se a execução for interrompida."""
    caminho = Path(caminho)
    temporario = caminho.with_suffix(caminho.suffix + ".tmp")
    temporario.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    temporario.replace(caminho)

def _texto_limpo(valor):
    if valor is None:
        return ""
    texto = str(valor).strip()
    return "" if texto.upper() in {"INDISPONIVEL", "NULL", "NONE"} else texto

def _sim_nao(valor):
    if valor is True:
        return "Sim"
    if valor is False:
        return "Não"
    return ""

def _status_captacao(valor):
    if valor is True:
        return "Aberto para captação"
    if valor is False:
        return "Fechado para captação"
    return ""

def _formatar_adiantamento(flag, modalidade, percentual):
    modalidade = _texto_limpo(modalidade)
    if flag is True:
        partes = ["Sim"]
        if modalidade and normalizar(modalidade) != "NAO SE APLICA":
            partes.append(modalidade.title())
        if percentual is not None and str(percentual).strip() != "":
            try:
                pct = float(percentual)
                partes.append(f"{pct:g}%")
            except (TypeError, ValueError):
                partes.append(str(percentual))
        return " · ".join(partes)
    if flag is False:
        return "Não se aplica" if normalizar(modalidade) == "NAO SE APLICA" else "Não disponível"
    return modalidade

def _inspecionar_campos_fundos_json(lista):
    """Debug: mostra campos do primeiro registro para identificar codfundo."""
    if not lista:
        return
    primeiro = lista[0]
    campos = list(primeiro.keys())
    log(f"[Fundos.json DEBUG] Campos do 1º registro ({len(campos)} campos):")
    for c in campos:
        val = primeiro.get(c)
        if val and str(val).strip():
            log(f"  {c!r:35s} = {str(val)[:60]!r}")

def _parsear_lista_fundos(lista):
    indice_exato = {}
    indice_palavras = []
    indice_codfundo = {}
    indice_cnpj_completo = {}
    mapa_cnpj = {}

    for f in lista:
        try:
            nome = _texto_limpo(f.get("no_fundo"))
            if not nome:
                continue

            url_raw = _texto_limpo(f.get("de_link_pagina_fundo"))
            codfundo = _extrair_codfundo(f)
            codigo_doc = codfundo.zfill(4) if codfundo and len(codfundo) < 4 else codfundo
            cnpj_raw = formatar_cnpj(f.get("nu_cnpj"))
            cnpj_limpo = re.sub(r'\D', '', cnpj_raw)

            # URLs oficiais do catálogo integral.
            doc_lamina = _texto_limpo(f.get("de_link_lamina"))
            doc_reg = _texto_limpo(f.get("de_link_regulamento"))
            doc_inf = _texto_limpo(f.get("de_link_info_compl"))
            doc_comunicado = _texto_limpo(f.get("de_link_fato_relevante"))
            doc_boletim = _texto_limpo(f.get("de_link_boletim_comercial"))
            doc_termo = _texto_limpo(f.get("de_link_termo_adesao"))
            doc_sumario = _texto_limpo(f.get("de_link_sumario"))
            doc_raio_x = _texto_limpo(f.get("de_link_raio_x"))

            # Somente documentos com padrão confiável recebem fallback.
            doc_carta = (
                f"https://www.caixa.gov.br/Downloads/aplicacao-financeira-carta-mensal/CM_{codfundo}.pdf"
                if codfundo else ""
            )
            if not _url_doc_valida(doc_lamina) and codigo_doc:
                doc_lamina = f"https://www.caixa.gov.br/downloads/aplicacao-financeira-laminas/LA_{codigo_doc}.pdf"
            if not _url_doc_valida(doc_reg) and codigo_doc:
                doc_reg = f"https://www.caixa.gov.br/downloads/aplicacao-financeira-regulamentos/RG_{codigo_doc}.pdf"
            if not _url_doc_valida(doc_inf) and codigo_doc:
                doc_inf = f"https://www.caixa.gov.br/Downloads/aplicacao-financeira-inf-com/FIC_{codigo_doc}.pdf"
            if not _url_doc_valida(doc_comunicado) and codfundo:
                doc_comunicado = f"https://www.caixa.gov.br/Downloads/aplicacao-financeira-comunicado-aos-cotistas/COM_{codfundo}.pdf"
            if not _url_doc_valida(doc_boletim) and codfundo:
                doc_boletim = f"https://www.caixa.gov.br/Downloads/aplicacao-financeira-laminas-comerciais/LAC_{codfundo}.pdf"

            dados = {
                "url": url_raw if _url_valida(url_raw) else "",
                "cnpj": cnpj_raw,
                "codfundo": codfundo,
                "perfil_risco": f.get("no_perfil_risco"),
                "taxa_adm": f.get("pc_taxa_adm_cliente"),
                "aplicacao_minima": f.get("vr_aplicacao_inicial"),
                "aplicacao_adicional_minima": f.get("vr_aplicacao_adicional_minima"),
                "resgate_minimo": f.get("vr_resgate_minimo"),
                "saldo_minimo": f.get("vr_saldo_minimo"),
                "conversao_aplicacao": f.get("de_conversao_aplicacao"),
                "conversao_resgate": f.get("de_conversao_resgate"),
                "pagamento_resgate": f.get("de_pagamento_resgate"),
                "benchmark": _texto_limpo(f.get("no_benchmark")),
                "estrategia": _texto_limpo(f.get("no_estrategia")),
                "captacao_aberta": f.get("ic_aberto_captacao"),
                "classificacao_tributaria": _texto_limpo(f.get("no_classificacao_tributaria")),
                "classificacao_investidor": _texto_limpo(f.get("no_classificacao_investidor")),
                "horario_limite": _texto_limpo(f.get("de_horario_limite")),
                "horario_resgate": _texto_limpo(f.get("de_horario_resgate")),
                "adiantamento_resgate": f.get("ic_adiantamento_resgate"),
                "adiantamento_modalidade": _texto_limpo(f.get("de_adiant_manual_automatico")),
                "adiantamento_percentual": f.get("pc_adiant_resgate"),
                "publico_alvo": f.get("lista_publico_alvo"),
                "segmentos": f.get("lista_segmento"),
                "movimentacao_automatica": f.get("ic_mov_automatica"),
                "carencia": f.get("ic_carencia"),
                "fim_carencia": _texto_limpo(f.get("dt_fim_carencia")),
                "asg": f.get("ic_asg"),
                "observacao_operacional": _texto_limpo(f.get("de_observacao_qs")),
                "razao_social": _texto_limpo(f.get("no_razao_social")),
                "docs": {
                    "lamina": doc_lamina if _url_doc_valida(doc_lamina) else "",
                    "regulamento": doc_reg if _url_doc_valida(doc_reg) else "",
                    "inf_comp": doc_inf if _url_doc_valida(doc_inf) else "",
                    "comunicado": doc_comunicado if _url_doc_valida(doc_comunicado) else "",
                    "carta_mensal": doc_carta,
                    "boletim": doc_boletim if _url_doc_valida(doc_boletim) else "",
                    "termo": doc_termo if _url_doc_valida(doc_termo) else "",
                    "sumario": doc_sumario if _url_doc_valida(doc_sumario) else "",
                    "raio_x": doc_raio_x if _url_doc_valida(doc_raio_x) else "",
                },
            }

            indice_exato[_normalizar_nome_fundo(nome)] = dados
            palavras = _palavras_chave_fundo(nome)
            if palavras:
                indice_palavras.append((frozenset(palavras), dados))

            if codfundo:
                chave_cod = str(codfundo).lstrip("0") or str(codfundo)
                indice_codfundo[chave_cod] = dados

            if cnpj_limpo:
                indice_cnpj_completo[cnpj_limpo] = dados
                mapa_cnpj[cnpj_limpo] = {
                    "codfundo": codfundo,
                    "nome": nome,
                    "cnpj": cnpj_raw,
                    "docs": dados["docs"],
                }
        except Exception as e:
            log(f"[Fundos.json] Registro ignorado por erro: {e}")
            continue

    return {
        "exato": indice_exato,
        "palavras": indice_palavras,
        "codfundo": indice_codfundo,
        "cnpj_completo": indice_cnpj_completo,
        "cnpj": mapa_cnpj,
    }

def _tokens_equivalentes(a, b):
    """Considera abreviações longas equivalentes sem relaxar demais o matching."""
    if a == b:
        return True
    if len(a) >= 5 and len(b) >= 5:
        return a.startswith(b) or b.startswith(a)
    return False

def _score_palavras_fundo(palavras_sipii, palavras_json):
    """Pontuação compatível com a regra antiga, com bônus seguro para abreviações."""
    if not palavras_sipii or not palavras_json:
        return 0.0
    usadas = set()
    equivalentes = 0
    for token_sipii in palavras_sipii:
        for token_json in palavras_json:
            if token_json in usadas:
                continue
            if _tokens_equivalentes(token_sipii, token_json):
                usadas.add(token_json)
                equivalentes += 1
                break
    return equivalentes / max(len(palavras_sipii), len(palavras_json))

def _buscar_meta_json(nome_sipii, indice_json):
    if not indice_json:
        return None

    chave = _normalizar_nome_fundo(nome_sipii)
    if chave in indice_json.get("exato", {}):
        return indice_json["exato"][chave]

    # Fallbacks estáticos são usados antes da aproximação porque apontam para
    # códigos oficiais confirmados e evitam falso positivo em nomes abreviados.
    cod_fb, razao_fb = _buscar_codfundo_por_nome(nome_sipii, indice_json)
    if cod_fb:
        chave_cod = str(cod_fb).lstrip("0") or str(cod_fb)
        meta_cod = indice_json.get("codfundo", {}).get(chave_cod)
        if meta_cod:
            log(f"  [Meta] Fallback nome→codfundo {cod_fb} para '{nome_sipii[:55]}' ({razao_fb})")
            return meta_cod

    palavras_sipii = _palavras_chave_fundo(nome_sipii)
    if not palavras_sipii:
        return None

    melhor_meta, melhor_score = None, 0.0
    for (palavras_json, dados) in indice_json.get("palavras", []):
        score = _score_palavras_fundo(palavras_sipii, palavras_json)
        if score > melhor_score:
            melhor_score = score
            melhor_meta = dados

    if melhor_score >= 0.65:
        return melhor_meta
    return None

def buscar_fundos_json(headers):
    lista = []
    origem = "nenhuma"
    log("[Fundos.json] Baixando catálogo integral da CAIXA Asset...")

    try:
        res = requests.get(FUNDOS_JSON_URL, headers=headers, timeout=35)
        if res.status_code == 200:
            raw = res.json()
            lista = raw if isinstance(raw, list) else (raw.get("fundos") or raw.get("value") or raw.get("data") or [])
            if not isinstance(lista, list):
                lista = []
            if lista:
                origem = "http"
                # O arquivo integral usado pela página é atualizado em toda execução bem-sucedida.
                _salvar_json_atomico(FUNDOS_JSON_LOCAL, lista)
                log(f"[Fundos.json] Download OK e arquivo local atualizado — {len(lista)} registros.")
        else:
            log(f"[Fundos.json] HTTP {res.status_code} — tentando arquivo local...")
    except Exception as e:
        log(f"[Fundos.json] Falha HTTP: {e} — tentando arquivo local...")

    if not lista and FUNDOS_JSON_LOCAL.exists():
        try:
            raw = json.loads(FUNDOS_JSON_LOCAL.read_text(encoding="utf-8"))
            lista = raw if isinstance(raw, list) else (raw.get("fundos") or raw.get("value") or raw.get("data") or [])
            if not isinstance(lista, list):
                lista = []
            if lista:
                origem = "local"
                log(f"[Fundos.json] FALLBACK local OK — {len(lista)} registros.")
        except Exception as e:
            log(f"[Fundos.json] Erro ao ler local: {e}")

    if not lista:
        log("[Fundos.json] Nenhuma fonte disponível. Continuando sem metadados.")
        return {}

    _inspecionar_campos_fundos_json(lista[:1])
    indice = _parsear_lista_fundos(lista)
    log(f"[Fundos.json] {len(indice['exato'])} produtos indexados.")

    mapa_cnpj = indice.get("cnpj", {})
    if mapa_cnpj:
        saida_html = {
            "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "origem": origem,
            "total": len(mapa_cnpj),
            "por_cnpj": mapa_cnpj,
        }
        try:
            _salvar_json_atomico(FUNDOS_CAIXA_PATH, saida_html)
            log(f"[Fundos.json] fundos_caixa.json salvo — {len(mapa_cnpj)} CNPJs indexados.")
        except Exception as e:
            log(f"[Fundos.json] Erro ao salvar fundos_caixa.json: {e}")

    indice["_origem"] = origem
    indice["_total_catalogo"] = len(lista)
    return indice

# ★ v17.2 — Mapeamento estático para fundos sem CNPJ no SIPII
# Chave: fragmento único e inequívoco do nome SIPII (lowercase, sem acento)
# Valor: codfundo confirmado no fundos_caixa.json
_CODFUNDO_FALLBACK = {
    "e-simples renda fixa":      "5980",   # CAIXA FIF E-SIMPLES RENDA FIXA LP
    "e simples renda fixa":      "5980",
    "e-simples fif rf":          "5980",   # CAIXA E-SIMPLES FIF RF LP
    "esimples fif rf":           "5980",
    "geracao jovem":             "5570",   # GER JOVEM RF CREDITO PRIV LP
    "geracao jov":               "5570",
    "exclusivo func":            "5972",   # EXCLUSIVO FUNCIONARIOS RF CRED PRIVADO
    "exclusivo funcionario":     "5972",
    "fifinvestidor":             "89",     # FIC FIF INVESTIDOR RF LONGO PRAZO
    "plus quali":                "6533",   # PLUS QUALIFICADO FIC FIF RF CRED PRIV
    "mega rf ref":               "5411",   # MEGA RF REF DI LONGO PRAZO
    "mega rf referenc":          "5411",
    "multimercado rv 30":        "82",     # FIF RV 30 MM LONGO PRAZO
    "etf ibovespa":              "5842",   # BRASIL ETF IBOVESPA FIF ACOES
    "ametista corp":             "7861",   # AMETISTA CORPORATIVO FIC FIF RF SIMPLES
    "ametista corporativo":      "7861",
    "transferencia voluntaria":  "5413",   # TRANSF VOLUNTARIAS POLIS FIC FIF RF CP
    "brasil idka ipca 2a":       "5825",   # BRASIL IDKA IPCA 2A TP FIF RF LP
    "idka ipca 2a":              "5825",
    "hashdex nasdaq crypto":     "7980",   # EXPERT HASHDEX NASDAQ CRYPTO INDEX MM
    "carteira quant quali":      "7991",   # CARTEIRA QUANTITATIVA QUALI MM LP
    "carteira quantitativa quali":"7991",
    # RS TITULOS PUBLICOS e BRASIL RF REFER DI LONGO: não estão na API CAIXA
}

def _buscar_codfundo_por_nome(nome_sipii, indice_json):
    """
    Fallback para fundos sem CNPJ: busca codfundo pelo nome.
    Etapa 1: mapeamento estático _CODFUNDO_FALLBACK (mais seguro).
    Etapa 2: não há match automático — retorna None (não arrisca falso positivo).
    """
    import unicodedata as _ud
    def _n(t):
        t = _ud.normalize('NFD', str(t))
        t = ''.join(c for c in t if _ud.category(c) != 'Mn')
        import re as _re
        return _re.sub(r'[^a-z0-9 ]', ' ', t.lower()).strip()

    nome_norm = _n(nome_sipii)
    # Etapa 1: mapeamento estático
    for frag, cod in _CODFUNDO_FALLBACK.items():
        if frag in nome_norm:
            return cod, f"fallback_estatico:{frag}"
    return None, None


def enriquecer_dados_com_fundos_json(df, indice_json):
    colunas_novas = [
        "CNPJ", "codfundo", "Perfil de Risco", "Taxa Adm (%)",
        "Aplicacao Minima (R$)", "Aplicacao Adicional Minima (R$)",
        "Resgate Minimo (R$)", "Saldo Minimo (R$)",
        "Conversao Aplicacao", "Conversao Resgate", "Pagamento Resgate",
        "Benchmark Oficial", "Estratégia", "Status de Captação",
        "Classificação Tributária", "Classificação Investidor",
        "Horário Limite Aplicação", "Horário Limite Resgate",
        "Adiantamento de Resgate", "Modalidade Adiantamento",
        "Percentual Adiantamento (%)", "Público Alvo", "Segmentos",
        "Movimentação Automática", "Carência", "Fim Carência", "ASG",
        "Razão Social", "Observação Operacional",
        "doc_lamina", "doc_regulamento", "doc_inf_comp", "doc_comunicado",
        "doc_carta", "doc_boletim", "doc_termo", "doc_sumario", "doc_raio_x",
    ]
    for col in colunas_novas:
        if col not in df.columns:
            df[col] = ""
    if not indice_json:
        return df

    vinculados = 0

    def _processar_linha(row):
        nonlocal vinculados
        url_atual = str(row.get("URL", "")).strip()
        meta = _buscar_meta_json(str(row.get("Fundo", "")), indice_json)
        if not meta:
            return row

        vinculados += 1
        if not _url_valida(url_atual):
            row["URL"] = meta.get("url", "")

        row["CNPJ"] = str(meta.get("cnpj") or "")
        row["codfundo"] = str(meta.get("codfundo") or "")
        row["Perfil de Risco"] = str(meta.get("perfil_risco") or "")
        row["Taxa Adm (%)"] = str(meta["taxa_adm"]) if meta.get("taxa_adm") is not None else ""
        row["Aplicacao Minima (R$)"] = str(meta["aplicacao_minima"]) if meta.get("aplicacao_minima") is not None else ""
        row["Aplicacao Adicional Minima (R$)"] = str(meta["aplicacao_adicional_minima"]) if meta.get("aplicacao_adicional_minima") is not None else ""
        row["Resgate Minimo (R$)"] = str(meta["resgate_minimo"]) if meta.get("resgate_minimo") is not None else ""
        row["Saldo Minimo (R$)"] = str(meta["saldo_minimo"]) if meta.get("saldo_minimo") is not None else ""
        row["Conversao Aplicacao"] = str(meta.get("conversao_aplicacao") or "")
        row["Conversao Resgate"] = str(meta.get("conversao_resgate") or "")
        row["Pagamento Resgate"] = str(meta.get("pagamento_resgate") or "")
        row["Benchmark Oficial"] = str(meta.get("benchmark") or "")
        row["Estratégia"] = str(meta.get("estrategia") or "").strip()
        row["Status de Captação"] = _status_captacao(meta.get("captacao_aberta"))
        row["Classificação Tributária"] = str(meta.get("classificacao_tributaria") or "")
        row["Classificação Investidor"] = str(meta.get("classificacao_investidor") or "")
        row["Horário Limite Aplicação"] = str(meta.get("horario_limite") or "")
        row["Horário Limite Resgate"] = str(meta.get("horario_resgate") or meta.get("horario_limite") or "")
        row["Adiantamento de Resgate"] = _formatar_adiantamento(
            meta.get("adiantamento_resgate"),
            meta.get("adiantamento_modalidade"),
            meta.get("adiantamento_percentual"),
        )
        row["Modalidade Adiantamento"] = str(meta.get("adiantamento_modalidade") or "")
        row["Percentual Adiantamento (%)"] = str(meta["adiantamento_percentual"]) if meta.get("adiantamento_percentual") is not None else ""
        row["Público Alvo"] = str(meta.get("publico_alvo") or "")
        row["Segmentos"] = str(meta.get("segmentos") or "")
        row["Movimentação Automática"] = _sim_nao(meta.get("movimentacao_automatica"))
        row["Carência"] = _sim_nao(meta.get("carencia"))
        row["Fim Carência"] = str(meta.get("fim_carencia") or "")
        row["ASG"] = _sim_nao(meta.get("asg"))
        row["Razão Social"] = str(meta.get("razao_social") or "")
        row["Observação Operacional"] = str(meta.get("observacao_operacional") or "")

        docs = meta.get("docs", {})
        row["doc_lamina"] = docs.get("lamina", "")
        row["doc_regulamento"] = docs.get("regulamento", "")
        row["doc_inf_comp"] = docs.get("inf_comp", "")
        row["doc_comunicado"] = docs.get("comunicado", "")
        row["doc_carta"] = docs.get("carta_mensal", "")
        row["doc_boletim"] = docs.get("boletim", "")
        row["doc_termo"] = docs.get("termo", "")
        row["doc_sumario"] = docs.get("sumario", "")
        row["doc_raio_x"] = docs.get("raio_x", "")
        return row

    df = df.apply(_processar_linha, axis=1)
    preenchidas = df["URL"].apply(_url_valida).sum()
    com_cod = (df.get("codfundo", pd.Series(dtype=str)).astype(str).str.strip() != "").sum()
    log(f"[Fundos.json] Vinculados: {vinculados}/{len(df)} | URLs válidas: {preenchidas}/{len(df)} | codfundo: {com_cod}/{len(df)}")

    auditoria = {
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "origem_fundos_json": indice_json.get("_origem", "desconhecida"),
        "registros_catalogo_oficial": indice_json.get("_total_catalogo", 0),
        "linhas_csv": int(len(df)),
        "vinculados": int(vinculados),
        "sem_vinculo": int(len(df) - vinculados),
        "com_cnpj": int((df["CNPJ"].astype(str).str.replace(r"\D", "", regex=True).str.len() == 14).sum()),
        "com_benchmark": int((df["Benchmark Oficial"].astype(str).str.strip() != "").sum()),
        "com_estrategia": int((df["Estratégia"].astype(str).str.strip() != "").sum()),
        "com_horario": int((df["Horário Limite Aplicação"].astype(str).str.strip() != "").sum()),
        "com_status_captacao": int((df["Status de Captação"].astype(str).str.strip() != "").sum()),
    }
    try:
        _salvar_json_atomico(FUNDOS_AUDITORIA_PATH, auditoria)
        log(f"[Auditoria] auditoria_fundos.json salvo — {vinculados}/{len(df)} vinculados.")
    except Exception as e:
        log(f"[Auditoria] Erro ao salvar auditoria_fundos.json: {e}")

    return df

# ---------------------------------------------------------------------------
# Fallback SIPII
# ---------------------------------------------------------------------------
def carregar_fallback_sipii():
    candidatos = [BASE_DIR / "dados_atuais.csv"]
    candidatos += sorted(BASE_DIR.glob("sipii_caixa_*.csv"), reverse=True)
    for caminho in candidatos:
        if caminho.exists():
            try:
                df = pd.read_csv(caminho, encoding="utf-8-sig")
                log(f"[FALLBACK SIPII] {len(df)} fundos de '{caminho.name}'.")
                if "Fundo_norm" not in df.columns:
                    df["Fundo_norm"] = df["Fundo"].apply(normalizar)
                return df
            except Exception as e:
                log(f"[FALLBACK SIPII] Erro em '{caminho.name}': {e}")
    log("[FALLBACK SIPII] Nenhum arquivo encontrado.")
    return pd.DataFrame()

# ---------------------------------------------------------------------------
# Limpeza de backups antigos
# ---------------------------------------------------------------------------
def limpar_backups_antigos(manter=5):
    for ext in ["csv", "xlsx"]:
        arquivos = sorted(BASE_DIR.glob(f"sipii_caixa_*.{ext}"), reverse=True)
        for arq in arquivos[manter:]:
            try:
                arq.unlink()
                log(f"[LIMPEZA] Removido: {arq.name}")
            except Exception as e:
                log(f"[LIMPEZA] Erro ao remover {arq.name}: {e}")

# ---------------------------------------------------------------------------
# KPIs do Dashboard — inalterado v15
# ---------------------------------------------------------------------------
def limpar_dados_para_calculo(df):
    df = df.copy()
    def converte_num(val):
        if pd.isna(val) or str(val).strip() in ['-', '—', '', 'None']: return None
        val = str(val).replace('"', '').strip()
        if ',' in val:
            if '.' in val and val.find('.') < val.find(','): val = val.replace('.', '')
            val = val.replace(',', '.')
        try: return float(val)
        except ValueError: return None
    for col in ['Cota (R$)','Variacao Dia (%)','Acum. Mes (%)','Acum. Ano (%)','Acum. 12M (%)','PL (milhoes R$)']:
        if col in df.columns:
            df[col] = df[col].apply(converte_num)
    return df

def gerar_json_kpis_dashboard(df_consolidado, caminho_saida):
    df_calculo = limpar_dados_para_calculo(df_consolidado)
    pl_total = df_calculo['PL (milhoes R$)'].sum()

    def w_avg(group):
        vals = group['Acum. 12M (%)']
        pesos = group['PL (milhoes R$)']
        mask = vals.notna() & pesos.notna() & (pesos > 0)
        if mask.sum() == 0: return None
        return (vals[mask] * pesos[mask]).sum() / pesos[mask].sum()

    categorias_kpi = {}
    for cat, group in df_calculo.groupby('Categoria'):
        wa = w_avg(group)
        categorias_kpi[cat] = {
            "qtd_ativos": int(group['Acum. 12M (%)'].notna().sum()),
            "pl_total": round(group['PL (milhoes R$)'].sum(), 2),
            "rent_12m_ponderada": round(wa, 2) if wa is not None else None,
        }

    perfil_kpi = {}
    for perf, group in df_calculo.groupby('Perfil'):
        pl_perf = group['PL (milhoes R$)'].sum()
        perfil_kpi[perf] = {
            "qtd_fundos": int(group['Fundo'].count()),
            "pl_total": round(pl_perf, 2),
            "share_percent": round((pl_perf / pl_total) * 100, 2) if pl_total > 0 else 0,
        }

    top_5 = df_calculo.nlargest(5, 'PL (milhoes R$)')
    kpis_finais = {
        "resumo_geral": {
            "pl_total_consolidado": round(pl_total, 2),
            "concentracao_top5_percent": round((top_5['PL (milhoes R$)'].sum() / pl_total) * 100, 2) if pl_total > 0 else 0,
            "pipeline_novos_fundos": int(df_calculo['Cota (R$)'].isna().sum()),
        },
        "categorias": categorias_kpi,
        "perfis_comerciais": perfil_kpi,
    }
    with open(caminho_saida, "w", encoding="utf-8") as f:
        json.dump(kpis_finais, f, indent=4, ensure_ascii=False)
    log(f"[KPIs] JSON exportado: {caminho_saida.name}")

# ---------------------------------------------------------------------------
# ★ Coletor de Indicadores de Mercado — v17.2 (CDI/IPCA 24M/36M + PTAX histórico 37M)
# ---------------------------------------------------------------------------
class ColetorMercado:
    def __init__(self):
        self.headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

    def _buscar_bcb(self, codigo_serie):
        try:
            url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo_serie}/dados/ultimos/1?formato=json"
            res = requests.get(url, headers=self.headers, timeout=10)
            if res.status_code == 200:
                return float(res.json()[0]['valor'])
        except Exception:
            pass
        return None

    # ──────────────────────────────────────────────────────────────────────
    # ★ v16 — CDI acumulado 12M / 24M / 36M via série 4391 (server-side)
    # Resolve o HTTP 400 que ocorre quando o browser tenta buscar esses dados
    # diretamente (api.bcb.gov.br não tem CORS confiável para browsers).
    # ──────────────────────────────────────────────────────────────────────
    def _buscar_cdi_acumulado(self):
        """
        CDI mensal com separação entre:
          - último mês fechado;
          - mês atual parcial, quando houver;
          - acumulado no ano fechado;
          - acumulado no ano com parcial;
          - acumulados 12M/24M/36M.

        A rotina aceita cdi_mensal_override.json em dois formatos:
        1) Legado/flat:
           {"2026-05": 1.07}

        2) Recomendado:
           {
             "confirmados": {"2026-05": 1.07},
             "parciais": {"2026-06": 0.03}
           }

        Valores confirmados prevalecem sobre a série BCB 4391 para evitar que
        parciais antigos sejam tratados como fechamento após a virada do mês.
        """
        log("[CDI] Calculando CDI mensal/ano/12M/24M/36M...")

        hoje = datetime.now()
        caminho_override = BASE_DIR / "cdi_mensal_override.json"

        confirmados = dict(FECHAMENTOS_CONFIRMADOS.get("cdi_mensal", {}))
        parciais = {}

        if caminho_override.exists():
            try:
                extra = json.loads(caminho_override.read_text(encoding="utf-8"))
                if isinstance(extra, dict) and ("confirmados" in extra or "parciais" in extra):
                    for k, v in (extra.get("confirmados") or {}).items():
                        try:
                            confirmados[str(k)] = float(str(v).replace(",", "."))
                        except Exception:
                            pass
                    for k, v in (extra.get("parciais") or {}).items():
                        try:
                            parciais[str(k)] = float(str(v).replace(",", "."))
                        except Exception:
                            pass
                    log(
                        f"  [CDI] Override local carregado: {caminho_override.name} "
                        f"({len(extra.get('confirmados') or {})} confirmados, {len(extra.get('parciais') or {})} parciais)"
                    )
                elif isinstance(extra, dict):
                    # Formato antigo: trata tudo como mês confirmado.
                    for k, v in extra.items():
                        try:
                            confirmados[str(k)] = float(str(v).replace(",", "."))
                        except Exception:
                            pass
                    log(f"  [CDI] Override local legado carregado: {caminho_override.name} ({len(extra)} registros)")
            except Exception as e:
                log(f"  [CDI] Erro ao ler cdi_mensal_override.json: {e}")

        def key_from_data_br(data_br):
            d, m, y = data_br.split("/")
            return f"{y}-{m.zfill(2)}"

        def label_from_key(key):
            y, m = key.split("-")
            return f"{MESES_PT[int(m)-1]}/{y}"

        def acum_valores(valores):
            f = 1.0
            for v in valores:
                try:
                    f *= (1 + float(v) / 100)
                except Exception:
                    pass
            return round((f - 1) * 100, 4)

        # Busca uma janela maior para conseguir 36M fechados.
        inicio = hoje - timedelta(days=42 * 31)
        url = (
            "https://api.bcb.gov.br/dados/serie/bcdata.sgs.4391/dados"
            f"?dataInicial={inicio.strftime('%d/%m/%Y')}"
            f"&dataFinal={hoje.strftime('%d/%m/%Y')}&formato=json"
        )

        serie = {}
        try:
            res = requests.get(url, headers=self.headers, timeout=25)
            if res.status_code == 200:
                for item in res.json():
                    try:
                        k = key_from_data_br(item["data"])
                        serie[k] = float(str(item["valor"]).replace(",", "."))
                    except Exception:
                        pass
                log(f"  [CDI] Série 4391 BCB carregada: {len(serie)} meses")
            else:
                log(f"  [CDI] Série 4391 BCB HTTP {res.status_code}")
        except Exception as e:
            log(f"  [CDI] Erro ao buscar série 4391 BCB: {e}")

        mes_atual_key = f"{hoje.year}-{hoje.month:02d}"
        primeiro_dia_mes = hoje.replace(day=1)
        mes_anterior_data = primeiro_dia_mes - timedelta(days=1)
        mes_fechado_key = f"{mes_anterior_data.year}-{mes_anterior_data.month:02d}"

        # Aplica fechamentos confirmados até o mês fechado.
        for k, v in confirmados.items():
            if k <= mes_fechado_key:
                serie[k] = float(v)
            elif k == mes_atual_key:
                # Se alguém informou o mês atual no arquivo legado, trata como parcial.
                parciais[k] = float(v)

        # Aplica parcial somente para o mês calendário atual.
        if mes_atual_key in parciais:
            serie[mes_atual_key] = float(parciais[mes_atual_key])

        if not serie:
            log("  [CDI] Sem série CDI disponível")
            return None

        # Todos os meses fechados até o mês anterior.
        chaves_fechadas = sorted(k for k in serie.keys() if k <= mes_fechado_key)
        mensal_fechado = serie.get(mes_fechado_key)
        parcial_mes_atual = serie.get(mes_atual_key) if mes_atual_key in parciais else None

        # Se, por algum motivo, não houver mês anterior na série, usa o último fechado disponível.
        if mensal_fechado is None and chaves_fechadas:
            mes_fechado_key = chaves_fechadas[-1]
            mensal_fechado = serie.get(mes_fechado_key)

        # Acumulado no ano fechado: jan até mês anterior.
        ano_atual = str(hoje.year)
        chaves_ano_fechado = sorted(k for k in chaves_fechadas if k.startswith(ano_atual + "-"))
        valores_ano_fechado = [serie[k] for k in chaves_ano_fechado]
        acum_ano = acum_valores(valores_ano_fechado) if valores_ano_fechado else None

        # Acumulado no ano com parcial: jan até mês atual, se houver parcial.
        chaves_ano_com_parcial = list(chaves_ano_fechado)
        if parcial_mes_atual is not None and mes_atual_key not in chaves_ano_com_parcial:
            chaves_ano_com_parcial.append(mes_atual_key)
        chaves_ano_com_parcial = sorted(chaves_ano_com_parcial)
        valores_ano_com_parcial = [serie[k] for k in chaves_ano_com_parcial if k in serie]
        acum_ano_com_parcial = acum_valores(valores_ano_com_parcial) if valores_ano_com_parcial else acum_ano

        # Acumulados de janela fechada.
        valores_fechados = [serie[k] for k in chaves_fechadas]
        acum_12m = acum_valores(valores_fechados[-12:]) if len(valores_fechados) >= 12 else None
        acum_24m = acum_valores(valores_fechados[-24:]) if len(valores_fechados) >= 24 else None
        acum_36m = acum_valores(valores_fechados[-36:]) if len(valores_fechados) >= 36 else None

        log(
            f"  [CDI] mês fechado={mensal_fechado}% ({label_from_key(mes_fechado_key)}) "
            f"| parcial={parcial_mes_atual}% ({label_from_key(mes_atual_key)}) "
            f"| ano fechado={acum_ano}% | ano c/ parcial={acum_ano_com_parcial}% "
            f"| 12M={acum_12m}% | 24M={acum_24m}% | 36M={acum_36m}%"
        )

        return {
            "mensal": mensal_fechado,
            "mes_ref": label_from_key(mes_fechado_key),
            "parcial_mes_atual": parcial_mes_atual,
            "parcial_ref": label_from_key(mes_atual_key),
            "acum_ano": acum_ano,
            "acum_ano_com_parcial": acum_ano_com_parcial,
            "acum_12m": acum_12m,
            "acum_24m": acum_24m,
            "acum_36m": acum_36m,
            "n_meses": len(chaves_fechadas),
            "fonte": "série 4391 BCB + fechamentos confirmados/override local",
            "status_mes_atual": "parcial" if parcial_mes_atual is not None else "sem_parcial",
            "historico": [{"key": k, "label": label_from_key(k), "valor": serie[k]} for k in sorted(serie.keys())],
        }

    # ──────────────────────────────────────────────────────────────────────
    # ★ v16 — PTAX histórico mensal (24M) — server-side
    # Resolve HTTP 400 do browser que tentava 36M (limite da API é 24M)
    # ──────────────────────────────────────────────────────────────────────
    # ──────────────────────────────────────────────────────────────────────
    # ★ v17.2 — PTAX histórico mensal robusto (37M)
    # Usa a API Olinda/CotacaoDolarPeriodo mês a mês e salva cache local.
    # Isso evita histórico vazio e mantém gráfico/cards/timeline na mesma base.
    # ──────────────────────────────────────────────────────────────────────
    def _buscar_ptax_historico(self, meses=37):
        """
        Retorna fechamentos mensais da PTAX de venda.

        Estrutura esperada pelo index.html:
        [
          {
            "key": "2026-05",
            "mes": "mai/2026",
            "cotacao": 5.0569,
            "var_pct": 1.37,
            "data_ref": "2026-05",
            "dataHoraCotacao": "2026-05-29T13:05:..."
          }
        ]

        Observações:
        - 37 meses = mês atual + 36 meses anteriores.
        - A busca é feita mês a mês para evitar limites/instabilidades do BCB.
        - Se a API falhar, tenta SGS 3697; se também falhar, usa cache local.
        """
        log(f"[PTAX] Buscando histórico {meses}M via CotacaoDolarPeriodo mês a mês...")

        hoje = datetime.now().date()

        def fmt_odata(d):
            # Formato exigido pela API Olinda PTAX: MM-DD-YYYY
            return d.strftime("%m-%d-%Y")

        def month_key(d):
            return f"{d.year}-{d.month:02d}"

        def last_day_of_month(year, month):
            return date(year, month, calendar.monthrange(year, month)[1])

        def add_months(d, qtd_meses):
            y = d.year + (d.month - 1 + qtd_meses) // 12
            m = (d.month - 1 + qtd_meses) % 12 + 1
            day = min(d.day, calendar.monthrange(y, m)[1])
            return date(y, m, day)

        def calc_var(atual, base):
            if atual is None or base is None or base == 0:
                return None
            return round((float(atual) / float(base) - 1) * 100, 2)

        def salvar_cache(historico):
            if not historico:
                return
            try:
                PTAX_CACHE_PATH.write_text(
                    json.dumps({
                        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                        "total": len(historico),
                        "ptax_historico": historico,
                    }, ensure_ascii=False, indent=2),
                    encoding="utf-8"
                )
                log(f"  [PTAX] Cache salvo: {PTAX_CACHE_PATH.name} ({len(historico)} meses)")
            except Exception as e:
                log(f"  [PTAX] Erro ao salvar cache: {e}")

        def carregar_cache():
            try:
                if not PTAX_CACHE_PATH.exists():
                    return []
                raw = json.loads(PTAX_CACHE_PATH.read_text(encoding="utf-8"))
                hist = raw.get("ptax_historico", []) if isinstance(raw, dict) else raw
                if hist:
                    log(f"  [PTAX] Usando cache local: {len(hist)} meses")
                    return hist[-meses:]
            except Exception as e:
                log(f"  [PTAX] Erro ao ler cache: {e}")
            return []

        def buscar_mes(inicio, fim):
            base_url = "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata"
            endpoint = (
                f"{base_url}/CotacaoDolarPeriodo"
                f"(dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)"
            )
            params = {
                "@dataInicial": f"'{fmt_odata(inicio)}'",
                "@dataFinalCotacao": f"'{fmt_odata(fim)}'",
                "$top": "100",
                "$format": "json",
                "$select": "cotacaoVenda,dataHoraCotacao",
            }

            for tentativa in range(1, 4):
                try:
                    res = requests.get(endpoint, params=params, headers=self.headers, timeout=25)
                    if res.status_code != 200:
                        log(f"  [PTAX] {month_key(inicio)} HTTP {res.status_code} — tentativa {tentativa}")
                        time.sleep(2 * tentativa)
                        continue

                    dados = res.json().get("value", [])
                    if not dados:
                        log(f"  [PTAX] {month_key(inicio)} sem dados")
                        return None

                    ultimo = max(dados, key=lambda x: str(x.get("dataHoraCotacao", "")))
                    cot = ultimo.get("cotacaoVenda")
                    data_hora = ultimo.get("dataHoraCotacao")
                    if cot is None or not data_hora:
                        return None

                    data_ref = str(data_hora)[:10]
                    dt_ref = datetime.strptime(data_ref, "%Y-%m-%d").date()
                    cotacao = round(float(str(cot).replace(",", ".")), 4)

                    return {
                        "key": month_key(dt_ref),
                        "mes": f"{MESES_PT[dt_ref.month - 1]}/{dt_ref.year}",
                        "cotacao": cotacao,
                        "var_pct": None,
                        "data_ref": data_ref,
                        "dataHoraCotacao": data_hora,
                    }
                except Exception as e:
                    log(f"  [PTAX] {month_key(inicio)} erro — tentativa {tentativa}: {e}")
                    time.sleep(2 * tentativa)
            return None

        primeiro_mes_atual = date(hoje.year, hoje.month, 1)
        inicio_janela = add_months(primeiro_mes_atual, -(meses - 1))
        registros = []

        for i in range(meses):
            ini = add_months(inicio_janela, i)
            fim = last_day_of_month(ini.year, ini.month)
            if fim > hoje:
                fim = hoje

            item = buscar_mes(ini, fim)
            if item:
                registros.append(item)
            time.sleep(0.25)

        # Remove duplicados e ordena por mês.
        por_mes = {}
        for item in registros:
            por_mes[item["key"]] = item
        historico = [por_mes[k] for k in sorted(por_mes.keys())]

        # Calcula variação mensal.
        prev = None
        for item in historico:
            item["var_pct"] = calc_var(item.get("cotacao"), prev) if prev else None
            prev = item.get("cotacao")

        if historico:
            log(f"  [PTAX] CotacaoDolarPeriodo OK — {len(historico)} meses")
            salvar_cache(historico)
            return historico[-meses:]

        # ── Fallback 1: SGS 3697 ──
        log("[PTAX] Olinda sem histórico. Tentando fallback SGS série 3697...")
        try:
            inicio_sgs = add_months(hoje, -(meses + 2))
            url_sgs = (
                "https://api.bcb.gov.br/dados/serie/bcdata.sgs.3697/dados"
                f"?dataInicial={inicio_sgs.strftime('%d/%m/%Y')}"
                f"&dataFinal={hoje.strftime('%d/%m/%Y')}&formato=json"
            )
            res = requests.get(url_sgs, headers=self.headers, timeout=25)
            if res.status_code == 200:
                dados = res.json()
                resultado = []
                prev = None
                for item in dados[-meses:]:
                    d, m, y = item["data"].split("/")
                    cotacao = round(float(str(item["valor"]).replace(",", ".")), 4)
                    resultado.append({
                        "key": f"{y}-{m}",
                        "mes": f"{MESES_PT[int(m)-1]}/{y}",
                        "cotacao": cotacao,
                        "var_pct": calc_var(cotacao, prev) if prev else None,
                        "data_ref": f"{y}-{m}-{d}",
                        "dataHoraCotacao": f"{y}-{m}-{d}T12:00:00",
                    })
                    prev = cotacao
                if resultado:
                    log(f"  [PTAX] SGS série 3697 OK — {len(resultado)} meses")
                    salvar_cache(resultado)
                    return resultado[-meses:]
            else:
                log(f"  [PTAX] SGS série 3697 HTTP {res.status_code}")
        except Exception as e:
            log(f"  [PTAX] SGS série 3697 erro: {e}")

        # ── Fallback 2: cache local ──
        cache = carregar_cache()
        if cache:
            return cache

        log("  [PTAX] Todas as estratégias falharam — PTAX histórico indisponível")
        return []

    def _montar_dolar_indice_ptax(self, ptax_historico):
        """
        Monta o bloco indices_mercado.dolar a partir do mesmo histórico PTAX
        usado no gráfico. Trata corretamente a virada do mês:
          - se ainda não houver cotação do mês atual, o último mês disponível
            vira mês fechado;
          - o campo variacao_mes_atual fica None, evitando exibir 0,00% no painel.
        """
        if not ptax_historico:
            return None

        hist = sorted(ptax_historico, key=lambda x: x.get("key", ""))
        hoje = datetime.now().date()
        mes_atual_key = f"{hoje.year}-{hoje.month:02d}"

        def calc_var(atual_val, base_val):
            if atual_val is None or base_val is None or base_val == 0:
                return None
            return round((float(atual_val) / float(base_val) - 1) * 100, 2)

        def item_n_meses_atras(n):
            return hist[-(n + 1)] if len(hist) > n else None

        # Se o último registro é do mês atual, existe mês parcial/atual.
        # Se o último registro é anterior ao mês atual, ele é o último mês fechado.
        latest = hist[-1]
        tem_mes_atual = latest.get("key") == mes_atual_key

        if tem_mes_atual:
            atual = latest
            mes_fechado = hist[-2] if len(hist) >= 2 else None
            mes_base = hist[-3] if len(hist) >= 3 else None
        else:
            atual = latest
            mes_fechado = latest
            mes_base = hist[-2] if len(hist) >= 2 else None

        ano_atual = int(str(atual.get("key", "0000-00"))[:4])
        dez_anterior = next((i for i in reversed(hist) if i.get("key") == f"{ano_atual - 1}-12"), None)
        base_12m = item_n_meses_atras(12)
        base_24m = item_n_meses_atras(24)
        base_36m = item_n_meses_atras(36)

        atual_val = atual.get("cotacao")
        mes_fechado_val = mes_fechado.get("cotacao") if mes_fechado else None
        mes_base_val = mes_base.get("cotacao") if mes_base else None

        variacao_mes_fechado = calc_var(mes_fechado_val, mes_base_val)
        override_ptax = FECHAMENTOS_CONFIRMADOS.get("ptax_mensal", {}).get(mes_fechado.get("key") if mes_fechado else "", {})
        if override_ptax.get("variacao_mes_fechado") is not None:
            variacao_mes_fechado = float(override_ptax["variacao_mes_fechado"])

        resultado = {
            "nome": "Dólar BRL/USD",
            "ticker": "PTAX",
            "fonte": "Banco Central do Brasil - PTAX",
            "mes_anterior_label": mes_fechado.get("mes") if mes_fechado else None,
            "mes_base_label": mes_base.get("mes") if mes_base else None,
            "fechamento_mes_anterior": round(mes_fechado_val, 4) if mes_fechado_val is not None else None,
            "data_mes_anterior": mes_fechado.get("data_ref") if mes_fechado else None,
            "fechamento_mes_base": round(mes_base_val, 4) if mes_base_val is not None else None,
            "data_mes_base": mes_base.get("data_ref") if mes_base else None,
            "variacao_mes_fechado": variacao_mes_fechado,
            "fechamento_atual": round(atual_val, 4) if atual_val is not None else None,
            "data_atual": atual.get("data_ref"),
            "variacao_mes_atual": calc_var(atual_val, mes_fechado_val) if tem_mes_atual else None,
            "acum_ano": calc_var(atual_val, dez_anterior.get("cotacao") if dez_anterior else None),
            "acum_12m": calc_var(atual_val, base_12m.get("cotacao") if base_12m else None),
            "acum_24m": calc_var(atual_val, base_24m.get("cotacao") if base_24m else None),
            "acum_36m": calc_var(atual_val, base_36m.get("cotacao") if base_36m else None),
            "base_ano": round(dez_anterior.get("cotacao"), 4) if dez_anterior else None,
            "base_12m": round(base_12m.get("cotacao"), 4) if base_12m else None,
            "base_24m": round(base_24m.get("cotacao"), 4) if base_24m else None,
            "base_36m": round(base_36m.get("cotacao"), 4) if base_36m else None,
            "tem_mes_atual": tem_mes_atual,
            "status_mes_atual": "parcial" if tem_mes_atual else "sem_parcial",
        }

        log(
            f"[PTAX] Dólar PTAX: último={resultado.get('fechamento_atual')} "
            f"({resultado.get('data_atual')}) | mês fechado={resultado.get('variacao_mes_fechado')}% "
            f"| mês atual={resultado.get('variacao_mes_atual')}% | ano={resultado.get('acum_ano')}% "
            f"| 12M={resultado.get('acum_12m')}% | 24M={resultado.get('acum_24m')}% | 36M={resultado.get('acum_36m')}%"
        )

        return resultado

    def _carregar_base_ipca(self):
        if IPCA_BASE_PATH.exists():
            try:
                dados = json.loads(IPCA_BASE_PATH.read_text(encoding="utf-8"))
                if dados:
                    log(f"[IPCA] Base local: {len(dados)} meses.")
                    return dados
            except Exception as e:
                log(f"[IPCA] Erro ao ler base local: {e}")
        log("[IPCA] Baixando série histórica SGS 433...")
        for tentativa in range(3):
            try:
                url = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.433/dados?formato=json"
                res = requests.get(url, headers=self.headers, timeout=60)
                if res.status_code != 200: continue
                historico = []
                for item in res.json():
                    d, m, y = item["data"].split("/")
                    historico.append({
                        "data": item["data"],
                        "label": f"{MESES_PT[int(m)-1]}/{y}",
                        "valor": round(float(item["valor"]), 4)
                    })
                log(f"[IPCA] Série completa: {len(historico)} meses.")
                return historico
            except Exception as e:
                log(f"[IPCA] Tentativa {tentativa+1} falhou: {e}")
                time.sleep(10 * (tentativa+1))
        return []

    def _buscar_ipca_delta(self, meses=3):
        for tentativa in range(3):
            try:
                url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.433/dados/ultimos/{meses}?formato=json"
                res = requests.get(url, headers=self.headers, timeout=20)
                if res.status_code == 200:
                    dados = res.json()
                    if dados: return dados
            except Exception as e:
                log(f"[IPCA] Delta tentativa {tentativa+1}: {e}")
                time.sleep(5 * (tentativa+1))
        return []

    def _merge_ipca(self, base, delta):
        """
        Consolida o histórico mensal do IPCA mantendo somente um registro por
        competência (AAAA-MM).

        Prioridade das fontes:
          1. base local sem override antigo;
          2. delta oficial da série 433 do BCB;
          3. override manual explicitamente configurado.

        A chave mensal evita que datas como 01/05/2026 e 31/05/2026 sejam
        tratadas como meses diferentes. Também corrige automaticamente bases
        antigas que já contenham duplicidades.
        """
        index = {}
        prioridades = {}

        def _normalizar_registro(data_br, valor, fonte_override=None):
            d, m, y = str(data_br).split("/")
            mes = int(m)
            ano = int(y)
            key = f"{ano:04d}-{mes:02d}"
            registro = {
                # A série 433 representa competência mensal; normalizamos para
                # o primeiro dia apenas para manter um formato único no JSON.
                "data": f"01/{mes:02d}/{ano:04d}",
                "key": key,
                "label": f"{MESES_PT[mes-1]}/{ano:04d}",
                "valor": round(float(valor), 4),
            }
            if fonte_override:
                registro["fonte_override"] = fonte_override
            return key, registro

        def _registrar(data_br, valor, prioridade, fonte_override=None):
            try:
                key, registro = _normalizar_registro(
                    data_br, valor, fonte_override
                )
            except (TypeError, ValueError, IndexError):
                return

            if key not in index or prioridade >= prioridades.get(key, -1):
                index[key] = registro
                prioridades[key] = prioridade

        # Base local. Registros antigos marcados como override recebem menor
        # prioridade do que o dado histórico normal, para limpar automaticamente
        # o caso mai/2026 0,50% versus o oficial 0,58%.
        for item in base or []:
            if not isinstance(item, dict):
                continue
            fonte_override = item.get("fonte_override")
            prioridade = 5 if fonte_override else 10
            _registrar(
                item.get("data"),
                item.get("valor"),
                prioridade,
                fonte_override,
            )

        # O delta oficial sempre substitui o valor já existente para o mês.
        for item in delta or []:
            if not isinstance(item, dict):
                continue
            _registrar(item.get("data"), item.get("valor"), 20)

        # Override manual é a última prioridade, mas somente quando estiver
        # expressamente configurado em FECHAMENTOS_CONFIRMADOS.
        for key, valor in FECHAMENTOS_CONFIRMADOS.get(
            "ipca_mensal", {}
        ).items():
            try:
                y, m = key.split("-")
                data_br = f"01/{int(m):02d}/{int(y):04d}"
                _registrar(
                    data_br,
                    valor,
                    30,
                    "fechamento confirmado",
                )
            except (TypeError, ValueError):
                log(f"[IPCA] Override mensal inválido ignorado: {key!r}")

        historico = [index[key] for key in sorted(index)]

        # Defesa final: nunca permitir mais de um item para a mesma competência.
        chaves = [item["key"] for item in historico]
        if len(chaves) != len(set(chaves)):
            raise RuntimeError("[IPCA] Falha ao eliminar competências duplicadas")

        return historico

    def _salvar_base_ipca(self, historico):
        try:
            IPCA_BASE_PATH.write_text(json.dumps(historico, ensure_ascii=False, indent=2), encoding="utf-8")
            log(f"[IPCA] Base salva: {len(historico)} meses.")
        except Exception as e:
            log(f"[IPCA] Erro ao salvar base: {e}")

    def _acumular(self, serie, n_meses):
        acc = 1.0
        for item in serie[-n_meses:]:
            try: acc *= (1 + float(item["valor"]) / 100)
            except: pass
        return round((acc - 1) * 100, 4)

    def _acumular_ano(self, serie):
        if not serie: return None
        ultimo_ano = serie[-1]["data"].split("/")[2]
        acc = 1.0
        for item in reversed(serie):
            if item["data"].split("/")[2] != ultimo_ano: break
            try: acc *= (1 + float(item["valor"]) / 100)
            except: pass
        return round((acc - 1) * 100, 4)

    def _buscar_yahoo(self, ticker):
        try:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?range=2mo&interval=1d"
            res = requests.get(url, headers=self.headers, timeout=10)
            if res.status_code == 200:
                quotes = res.json()['chart']['result'][0]['indicators']['quote'][0]['close']
                quotes = [q for q in quotes if q is not None]
                if quotes:
                    return {"atual": quotes[-1], "anterior": quotes[-22] if len(quotes) >= 22 else quotes[0]}
        except Exception:
            pass
        return {"atual": None, "anterior": None}

    # ──────────────────────────────────────────────────────────────────────
    # v17 — Índices de mercado com mês fechado, mês atual, ano, 12M, 24M e 36M
    # ──────────────────────────────────────────────────────────────────────
    def _variacao_pct(self, atual, anterior):
        try:
            if atual is None or anterior is None:
                return None
            atual = float(atual)
            anterior = float(anterior)
            if anterior == 0:
                return None
            return round(((atual / anterior) - 1) * 100, 2)
        except Exception:
            return None

    def _buscar_yahoo_historico(self, ticker, dias=1250):
        """
        Busca histórico diário no Yahoo Finance.
        1250 dias cobre aproximadamente 36 meses com folga.
        """
        fim = datetime.now() + timedelta(days=1)
        inicio = fim - timedelta(days=dias)

        period1 = int(inicio.timestamp())
        period2 = int(fim.timestamp())

        url = (
            f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
            f"?period1={period1}&period2={period2}&interval=1d"
        )

        try:
            log(f"[Yahoo Histórico] Buscando {ticker}...")
            res = requests.get(url, headers=self.headers, timeout=25)

            if res.status_code != 200:
                log(f"[Yahoo Histórico] {ticker} HTTP {res.status_code}")
                return pd.DataFrame()

            data = res.json()
            result = data.get("chart", {}).get("result", [])

            if not result:
                log(f"[Yahoo Histórico] {ticker} sem resultado")
                return pd.DataFrame()

            result = result[0]
            timestamps = result.get("timestamp", [])
            quote = result.get("indicators", {}).get("quote", [{}])[0]
            closes = quote.get("close", [])

            linhas = []
            for ts, close in zip(timestamps, closes):
                if close is None:
                    continue

                dt = datetime.fromtimestamp(ts).date()
                linhas.append({
                    "data": pd.to_datetime(dt),
                    "fechamento": float(close),
                })

            df = pd.DataFrame(linhas)

            if df.empty:
                log(f"[Yahoo Histórico] {ticker} retornou vazio")
                return df

            df = df.sort_values("data").drop_duplicates("data")
            df["ano_mes"] = df["data"].dt.strftime("%Y-%m")
            log(f"[Yahoo Histórico] {ticker} OK — {len(df)} registros")
            return df

        except Exception as e:
            log(f"[Yahoo Histórico] Erro em {ticker}: {e}")
            return pd.DataFrame()

    def _ultimo_registro_ate(self, df, data_limite):
        """
        Retorna o último fechamento disponível até determinada data.
        Resolve fins de semana e feriados.
        """
        if df.empty:
            return None

        data_limite = pd.to_datetime(data_limite)
        filtrado = df[df["data"] <= data_limite]

        if filtrado.empty:
            return None

        return filtrado.iloc[-1].to_dict()

    def _fechamento_final_mes(self, df, ano_mes):
        """
        Retorna o último fechamento disponível dentro de um mês calendário.
        Exemplo: 2026-04 retorna o último pregão de abril.
        """
        if df.empty:
            return None

        filtrado = df[df["ano_mes"] == ano_mes]

        if filtrado.empty:
            return None

        return filtrado.iloc[-1].to_dict()

    def _datas_referencia_indices(self):
        """
        Define as referências:
        - mês anterior fechado;
        - mês base anterior ao mês fechado;
        - início do ano;
        - bases de 12M, 24M e 36M.
        """
        hoje = pd.Timestamp(datetime.now().date())

        primeiro_mes_atual = hoje.replace(day=1)
        ultimo_dia_mes_anterior = primeiro_mes_atual - pd.Timedelta(days=1)
        primeiro_mes_anterior = ultimo_dia_mes_anterior.replace(day=1)
        ultimo_dia_mes_base = primeiro_mes_anterior - pd.Timedelta(days=1)

        mes_anterior_key = ultimo_dia_mes_anterior.strftime("%Y-%m")
        mes_base_key = ultimo_dia_mes_base.strftime("%Y-%m")

        mes_anterior_label = f"{MESES_PT[ultimo_dia_mes_anterior.month - 1]}/{ultimo_dia_mes_anterior.year}"
        mes_base_label = f"{MESES_PT[ultimo_dia_mes_base.month - 1]}/{ultimo_dia_mes_base.year}"

        return {
            "hoje": hoje,
            "mes_anterior_key": mes_anterior_key,
            "mes_base_key": mes_base_key,
            "mes_anterior_label": mes_anterior_label,
            "mes_base_label": mes_base_label,
            "data_base_ano": pd.Timestamp(year=hoje.year - 1, month=12, day=31),
            "data_base_12m": hoje - pd.DateOffset(months=12),
            "data_base_24m": hoje - pd.DateOffset(months=24),
            "data_base_36m": hoje - pd.DateOffset(months=36),
        }

    def _resumo_indice_yahoo(self, ticker, nome, dias=1250):
        """
        Calcula fechamento mensal, mês atual, ano, 12M, 24M e 36M.
        Trata corretamente a virada do mês: se ainda não houver registro do mês
        atual, o último pregão disponível é tratado como mês fechado, não como
        mês atual 0,00%.
        """
        log(f"[Índices] Calculando {nome} ({ticker})...")

        df = self._buscar_yahoo_historico(ticker, dias=dias)
        ref = self._datas_referencia_indices()

        if df.empty:
            return {
                "nome": nome,
                "ticker": ticker,
                "fonte": "Yahoo Finance",
                "erro": "sem dados",
                "mes_anterior_label": ref["mes_anterior_label"],
                "mes_base_label": ref["mes_base_label"],
            }

        fechamento_mes_anterior = self._fechamento_final_mes(df, ref["mes_anterior_key"])
        fechamento_mes_base = self._fechamento_final_mes(df, ref["mes_base_key"])
        fechamento_atual = df.iloc[-1].to_dict()

        # Se o último dado ainda pertence ao mês anterior, não existe parcial do mês atual.
        ultimo_ano_mes = str(fechamento_atual.get("data"))[:7] if fechamento_atual else ""
        tem_mes_atual = ultimo_ano_mes == datetime.now().strftime("%Y-%m")

        fechamento_base_ano = self._ultimo_registro_ate(df, ref["data_base_ano"])
        fechamento_base_12m = self._ultimo_registro_ate(df, ref["data_base_12m"])
        fechamento_base_24m = self._ultimo_registro_ate(df, ref["data_base_24m"])
        fechamento_base_36m = self._ultimo_registro_ate(df, ref["data_base_36m"])

        atual = fechamento_atual.get("fechamento") if fechamento_atual else None
        mes_ant = fechamento_mes_anterior.get("fechamento") if fechamento_mes_anterior else None
        mes_base = fechamento_mes_base.get("fechamento") if fechamento_mes_base else None

        base_ano = fechamento_base_ano.get("fechamento") if fechamento_base_ano else None
        base_12m = fechamento_base_12m.get("fechamento") if fechamento_base_12m else None
        base_24m = fechamento_base_24m.get("fechamento") if fechamento_base_24m else None
        base_36m = fechamento_base_36m.get("fechamento") if fechamento_base_36m else None

        variacao_mes_fechado = self._variacao_pct(mes_ant, mes_base)
        overrides_ticker = FECHAMENTOS_CONFIRMADOS.get("indices_mensais", {}).get(ticker, {})
        override_mes = overrides_ticker.get(ref["mes_anterior_key"], {})
        if override_mes.get("fechamento") is not None:
            mes_ant = float(override_mes["fechamento"])
        if override_mes.get("variacao_mes_fechado") is not None:
            variacao_mes_fechado = float(override_mes["variacao_mes_fechado"])

        resultado = {
            "nome": nome,
            "ticker": ticker,
            "fonte": "Yahoo Finance",

            "mes_anterior_label": ref["mes_anterior_label"],
            "mes_base_label": ref["mes_base_label"],

            "fechamento_mes_anterior": round(mes_ant, 2) if mes_ant is not None else None,
            "data_mes_anterior": fechamento_mes_anterior["data"].strftime("%Y-%m-%d") if fechamento_mes_anterior else None,

            "fechamento_mes_base": round(mes_base, 2) if mes_base is not None else None,
            "data_mes_base": fechamento_mes_base["data"].strftime("%Y-%m-%d") if fechamento_mes_base else None,

            "variacao_mes_fechado": variacao_mes_fechado,

            "fechamento_atual": round(atual, 2) if atual is not None else None,
            "data_atual": fechamento_atual["data"].strftime("%Y-%m-%d") if fechamento_atual else None,

            "variacao_mes_atual": self._variacao_pct(atual, mes_ant) if tem_mes_atual else None,
            "acum_ano": self._variacao_pct(atual, base_ano),
            "acum_12m": self._variacao_pct(atual, base_12m),
            "acum_24m": self._variacao_pct(atual, base_24m),
            "acum_36m": self._variacao_pct(atual, base_36m),

            "base_ano": round(base_ano, 2) if base_ano is not None else None,
            "base_12m": round(base_12m, 2) if base_12m is not None else None,
            "base_24m": round(base_24m, 2) if base_24m is not None else None,
            "base_36m": round(base_36m, 2) if base_36m is not None else None,
            "tem_mes_atual": tem_mes_atual,
            "status_mes_atual": "parcial" if tem_mes_atual else "sem_parcial",
        }

        log(
            f"[Índices] {nome}: "
            f"mês fechado={resultado.get('variacao_mes_fechado')}% | "
            f"mês atual={resultado.get('variacao_mes_atual')}% | "
            f"ano={resultado.get('acum_ano')}% | "
            f"12M={resultado.get('acum_12m')}% | "
            f"24M={resultado.get('acum_24m')}% | "
            f"36M={resultado.get('acum_36m')}%"
        )

        return resultado

    def _converter_indice_usd_para_brl(self, indice_usd, dolar_ref):
        """
        Converte índices americanos para BRL usando PTAX.
        Respeita o status de virada de mês: sem parcial do mês atual, a variação
        do mês atual em BRL fica None e o mês fechado segue em destaque.
        """
        if not indice_usd or not dolar_ref:
            return indice_usd

        try:
            fechamento_atual_usd = indice_usd.get("fechamento_atual")
            fechamento_mes_ant_usd = indice_usd.get("fechamento_mes_anterior")
            fechamento_mes_base_usd = indice_usd.get("fechamento_mes_base")

            base_ano_usd = indice_usd.get("base_ano")
            base_12m_usd = indice_usd.get("base_12m")
            base_24m_usd = indice_usd.get("base_24m")
            base_36m_usd = indice_usd.get("base_36m")

            dolar_atual = dolar_ref.get("fechamento_atual")
            dolar_mes_ant = dolar_ref.get("fechamento_mes_anterior")
            dolar_mes_base = dolar_ref.get("fechamento_mes_base")

            dolar_base_ano = dolar_ref.get("base_ano")
            dolar_base_12m = dolar_ref.get("base_12m")
            dolar_base_24m = dolar_ref.get("base_24m")
            dolar_base_36m = dolar_ref.get("base_36m")

            atual_brl = fechamento_atual_usd * dolar_atual if fechamento_atual_usd and dolar_atual else None
            mes_ant_brl = fechamento_mes_ant_usd * dolar_mes_ant if fechamento_mes_ant_usd and dolar_mes_ant else None
            mes_base_brl = fechamento_mes_base_usd * dolar_mes_base if fechamento_mes_base_usd and dolar_mes_base else None

            base_ano_brl = base_ano_usd * dolar_base_ano if base_ano_usd and dolar_base_ano else None
            base_12m_brl = base_12m_usd * dolar_base_12m if base_12m_usd and dolar_base_12m else None
            base_24m_brl = base_24m_usd * dolar_base_24m if base_24m_usd and dolar_base_24m else None
            base_36m_brl = base_36m_usd * dolar_base_36m if base_36m_usd and dolar_base_36m else None

            indice_usd["fechamento_atual_brl"] = round(atual_brl, 2) if atual_brl else None
            indice_usd["fechamento_mes_anterior_brl"] = round(mes_ant_brl, 2) if mes_ant_brl else None
            indice_usd["fechamento_mes_base_brl"] = round(mes_base_brl, 2) if mes_base_brl else None

            indice_usd["base_ano_brl"] = round(base_ano_brl, 2) if base_ano_brl else None
            indice_usd["base_12m_brl"] = round(base_12m_brl, 2) if base_12m_brl else None
            indice_usd["base_24m_brl"] = round(base_24m_brl, 2) if base_24m_brl else None
            indice_usd["base_36m_brl"] = round(base_36m_brl, 2) if base_36m_brl else None

            var_fechado_brl = self._variacao_pct(mes_ant_brl, mes_base_brl)
            override_mes = FECHAMENTOS_CONFIRMADOS.get("indices_mensais", {}).get(
                indice_usd.get("ticker"), {}
            ).get(datetime.now().replace(day=1).strftime("%Y-%m"), {})

            # O override de BRL normalmente será informado para o mês anterior fechado,
            # não para o mês atual. Corrige a chave para o campo mes_anterior_key.
            mes_anterior_key = None
            try:
                ref = self._datas_referencia_indices()
                mes_anterior_key = ref.get("mes_anterior_key")
            except Exception:
                pass
            if mes_anterior_key:
                override_mes = FECHAMENTOS_CONFIRMADOS.get("indices_mensais", {}).get(
                    indice_usd.get("ticker"), {}
                ).get(mes_anterior_key, {})

            if override_mes.get("variacao_mes_fechado_brl") is not None:
                var_fechado_brl = float(override_mes["variacao_mes_fechado_brl"])

            indice_usd["variacao_mes_fechado_brl"] = var_fechado_brl
            indice_usd["variacao_mes_atual_brl"] = self._variacao_pct(atual_brl, mes_ant_brl) if indice_usd.get("tem_mes_atual") else None
            indice_usd["acum_ano_brl"] = self._variacao_pct(atual_brl, base_ano_brl)
            indice_usd["acum_12m_brl"] = self._variacao_pct(atual_brl, base_12m_brl)
            indice_usd["acum_24m_brl"] = self._variacao_pct(atual_brl, base_24m_brl)
            indice_usd["acum_36m_brl"] = self._variacao_pct(atual_brl, base_36m_brl)

        except Exception as e:
            log(f"[Índices BRL] Erro ao converter {indice_usd.get('nome')}: {e}")

        return indice_usd

    def _carregar_base_selic(self):
        if SELIC_BASE_PATH.exists():
            try:
                raw = json.loads(SELIC_BASE_PATH.read_text(encoding="utf-8"))
                historico = []
                for item in raw.get("conteudo", []):
                    data_iso = item.get("DataReuniaoCopom", "").split("T")[0]
                    dt = datetime.strptime(data_iso, "%Y-%m-%d") if data_iso else None
                    historico.append({
                        "data": dt.strftime("%d/%m/%Y") if dt else "-",
                        "valor": item.get("MetaSelic"),
                        "numero_reuniao": item.get("NumeroReuniaoCopom"),
                        "DataReuniaoCopom": item.get("DataReuniaoCopom"),
                        "MetaSelic": item.get("MetaSelic"),
                    })
                return historico
            except Exception as e:
                log(f"[SELIC] Erro ao ler histórico local: {e}")
        return []

    def _carregar_base_meta_inflacao(self):
        if META_INFLACAO_PATH.exists():
            try:
                raw = json.loads(META_INFLACAO_PATH.read_text(encoding="utf-8"))
                return raw.get("conteudo", [])
            except Exception as e:
                log(f"[META INFLAÇÃO] Erro: {e}")
        return []


    def _sincronizar_meta_inflacao(self, serie, label_mes, valor_12m):
        """Insere/substitui na série do gráfico o IPCA 12M mais recente.

        A base meta-vs-inflacao-efetiva pode ser publicada com atraso em relação
        à série mensal 433. Como o acumulado de 12 meses já foi calculado com os
        dados oficiais, esta rotina garante que o gráfico termine na mesma
        competência e no mesmo valor exibido pelo card.
        """
        if valor_12m is None or not label_mes:
            return serie or []

        mapa_meses = {nome: i + 1 for i, nome in enumerate(MESES_PT)}
        try:
            mes_txt, ano_txt = str(label_mes).strip().lower().split("/")
            mes = mapa_meses[mes_txt]
            ano = int(ano_txt)
            chave_atual = f"{ano:04d}-{mes:02d}"
        except (ValueError, KeyError):
            log(f"[META INFLAÇÃO] Competência inválida: {label_mes!r}")
            return serie or []

        por_mes = {}
        for item in serie or []:
            if not isinstance(item, dict):
                continue
            data_ref = str(item.get("DataReferencia") or "")
            chave = data_ref[:7]
            try:
                valor = float(item.get("Inflacao12Meses"))
            except (TypeError, ValueError):
                continue
            if re.fullmatch(r"\d{4}-\d{2}", chave):
                registro = dict(item)
                registro["Inflacao12Meses"] = valor
                por_mes[chave] = registro

        anterior = por_mes.get(chave_atual, {})
        por_mes[chave_atual] = {
            **anterior,
            "DataReferencia": f"{chave_atual}-01T03:00:00Z",
            "Inflacao12Meses": round(float(valor_12m), 4),
            "CartaAberta": anterior.get("CartaAberta", "Não"),
        }

        resultado = [por_mes[chave] for chave in sorted(por_mes, reverse=True)]
        log(
            f"  [META INFLAÇÃO] gráfico sincronizado: {label_mes} = "
            f"{round(float(valor_12m), 4)}% em 12 meses."
        )
        return resultado

    # ──────────────────────────────────────────────────────────────────────
    # ★ v17.1 — Poupança nova + antiga: mensal atual + acum. ano
    #
    # Série BCB usadas:
    #   196 — Poupança nova (depósitos a partir de 04/05/2012) — mensal
    #   253  — TR mensal (só quando Selic ≤ 8,5%; hoje não se aplica)
    #
    # Regra:
    #   Nova:  Selic > 8,5% → TR + 0,50% a.m.
    #          Selic ≤ 8,5% → TR + 70% × (Selic/12)
    #   Antiga: sempre TR + 0,50% a.m. (independente da Selic)
    #
    # Quando Selic > 8,5% (cenário atual), nova = antiga em valor mensal.
    # ──────────────────────────────────────────────────────────────────────
    def _buscar_poupanca_detalhada(self, selic_meta):
        """
        Retorna dicionário com:
          nova:  { valor, mensal, acum_ano, historico_ano, nota }
          antiga:{ valor, mensal, acum_ano, historico_ano, nota }

        Lógica das regras da poupança:
          Selic > 8,5% a.a. → NOVA = TR + 0,50% a.m.  |  ANTIGA = TR + 0,50% a.m.
                               → ambas têm a MESMA taxa: nova == antiga
          Selic ≤ 8,5% a.a. → NOVA = TR + 70% × (Selic/12)  |  ANTIGA = TR + 0,50% a.m.

        Para a nova: usa série BCB 196 (poupança nova mensal) — valor real e confiável.
        Para a antiga quando Selic > 8,5%: COPIA os dados da nova (matematicamente idêntico).
        Para a antiga quando Selic ≤ 8,5%: usa série BCB 253 (TR mensal) + 0,50%.
        NÃO usa série 7814 — ela retorna taxas anuais (~CDI), não TR mensal.
        """
        hoje    = datetime.now()
        ini_ano = datetime(hoje.year, 1, 1)
        fmt_d   = lambda d: d.strftime('%d/%m/%Y')
        acima   = selic_meta is not None and selic_meta > 8.5

        def _fetch_serie_ano(serie):
            """Busca série BCB do início do ano até hoje; retorna lista [{data, valor}]."""
            url = (
                f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{serie}/dados"
                f"?dataInicial={fmt_d(ini_ano)}&dataFinal={fmt_d(hoje)}&formato=json"
            )
            try:
                res = requests.get(url, headers=self.headers, timeout=15)
                if res.status_code == 200:
                    dados = res.json()
                    if dados:
                        return [
                            {"data": item["data"],
                             "valor": round(float(str(item["valor"]).replace(",",".")), 6)}
                            for item in dados
                        ]
            except Exception as e:
                log(f"  [Poupança] Erro série {serie}: {e}")
            return []

        def _acumular(lista):
            """Composição: (1+r1/100) × (1+r2/100) × … − 1, em %."""
            acc = 1.0
            for item in lista:
                try:
                    acc *= (1 + item["valor"] / 100)
                except Exception:
                    pass
            return round((acc - 1) * 100, 4) if lista else None

        def _validar_mensal(lista, nome):
            """Descarta lista se valores > 5% a.m. (indica série errada)."""
            if not lista:
                return lista
            if any(abs(item["valor"]) > 5.0 for item in lista):
                log(f"  [Poupança] AVISO: {nome} com valores > 5%/mes — serie incorreta, descartando")
                return []
            return lista

        log("[Poupança] Buscando série 196 (nova) para acum. ano...")

        # ── Poupança nova (série 196 — mais confiável, específica para nova) ──
        dados_nova = _validar_mensal(_fetch_serie_ano(196), "nova (196)")
        mensal_nova = dados_nova[-1]["valor"] if dados_nova else None
        acum_nova   = _acumular(dados_nova)
        log(f"  [Poupança] nova: mensal={mensal_nova}% | acum. ano={acum_nova}% ({len(dados_nova)} meses)")

        # ── Poupança antiga ────────────────────────────────────────────────────
        if acima:
            # Selic > 8,5%: antiga = nova (mesma fórmula: TR + 0,50% a.m.)
            # Não precisamos buscar TR separada — os valores são idênticos.
            dados_antiga  = [{"data": item["data"], "valor": item["valor"]}
                             for item in dados_nova]
            mensal_antiga = mensal_nova
            acum_antiga   = acum_nova
            log(f"  [Poupança] antiga = nova (Selic {selic_meta}% > 8,5% → TR+0,50% para ambas)")
        else:
            # Selic ≤ 8,5%: antiga = TR + 0,50% (série 253 — TR mensal)
            # Nova usa 70% da Selic/12, então as taxas divergem.
            log("  [Poupança] Selic <= 8,5%: buscando TR (série 253) para poupança antiga...")
            dados_tr = _validar_mensal(_fetch_serie_ano(253), "TR (253)")
            if dados_tr:
                dados_antiga = [{"data": item["data"],
                                 "valor": round(item["valor"] + 0.50, 6)}
                                for item in dados_tr]
                mensal_antiga = dados_antiga[-1]["valor"] if dados_antiga else None
                acum_antiga   = _acumular(dados_antiga)
                log(f"  [Poupança] antiga via TR+0,50%: mensal={mensal_antiga}% | acum={acum_antiga}%")
            else:
                # Fallback conservador: estima 0,50% a.m. (TR ≈ 0 quando Selic ≤ 8,5%)
                dados_antiga  = []
                mensal_antiga = 0.50
                acum_antiga   = None
                log("  [Poupança] antiga: TR indisponível — usando 0,50% a.m. como estimativa")

        nota_nova = (
            "TR + 0,50% a.m. (Selic > 8,5% a.a.)" if acima
            else "TR + 70% da Selic a.m. (Selic <= 8,5% a.a.)"
        )

        return {
            "nova": {
                "valor":         mensal_nova,
                "mensal":        mensal_nova,
                "acum_ano":      acum_nova,
                "historico_ano": dados_nova,
                "unidade":       "% a.m.",
                "nota":          nota_nova,
            },
            "antiga": {
                "valor":         mensal_antiga,
                "mensal":        mensal_antiga,
                "acum_ano":      acum_antiga,
                "historico_ano": dados_antiga,
                "unidade":       "% a.m.",
                "nota":          "TR + 0,50% a.m. — independente da Selic",
            },
        }

    def coletar_todos(self):
        log("[MERCADO] Coletando indicadores macro (v17 — índices mercado ampliados)...")

        # Taxas de referência
        selic_meta = self._buscar_bcb(432)

        # ★ v17.1 — Poupança nova + antiga com acum. ano (série 196; 253 para Selic≤8,5%)
        poup_detalhado = self._buscar_poupanca_detalhada(selic_meta)
        poupanca_nova  = poup_detalhado["nova"]["valor"] if poup_detalhado else self._buscar_bcb(196)

        # CDI acumulado server-side (resolve CORS/400 no browser)
        cdi_acum = self._buscar_cdi_acumulado()
        cdi_mensal_atual = cdi_acum["mensal"] if cdi_acum else self._buscar_bcb(4391)

        # IPCA histórico
        selic_historico       = self._carregar_base_selic()
        inflacao_meta_efetiva = self._carregar_base_meta_inflacao()
        base_ipca  = self._carregar_base_ipca()
        delta_ipca = self._buscar_ipca_delta(meses=3)
        ipca_serie = self._merge_ipca(base_ipca, delta_ipca)
        if ipca_serie:
            self._salvar_base_ipca(ipca_serie)

        ipca_ultimo_mes = ipca_label_mes = ipca_acum_ano = ipca_acum_12m = None
        ipca_acum_24m = ipca_acum_36m = None
        ipca_historico = []
        if ipca_serie:
            ultimo = ipca_serie[-1]
            ipca_ultimo_mes = ultimo["valor"]
            ipca_label_mes  = ultimo["label"]
            ipca_acum_ano   = self._acumular_ano(ipca_serie)
            ipca_acum_12m   = self._acumular(ipca_serie, 12)
            ipca_acum_24m   = self._acumular(ipca_serie, 24) if len(ipca_serie) >= 24 else None
            ipca_acum_36m   = self._acumular(ipca_serie, 36) if len(ipca_serie) >= 36 else None
            ipca_historico  = [{"label": i["label"], "valor": i["valor"]} for i in ipca_serie]
            log(f"  [IPCA] 12M={ipca_acum_12m}% | 24M={ipca_acum_24m}% | 36M={ipca_acum_36m}%")

        # Mantém o gráfico IPCA 12M × meta na mesma competência do card.
        inflacao_meta_efetiva = self._sincronizar_meta_inflacao(
            inflacao_meta_efetiva,
            ipca_label_mes,
            ipca_acum_12m,
        )

        # PTAX histórico mensal — base única para card, tabela, timeline e gráfico do dólar.
        # 37 meses = mês atual + 36 meses anteriores.
        ptax_historico = self._buscar_ptax_historico(meses=37)
        dolar_indice = self._montar_dolar_indice_ptax(ptax_historico)

        # Fallback: se PTAX falhar totalmente, usa Yahoo apenas para não deixar o painel vazio.
        if not dolar_indice:
            dolar_indice = self._resumo_indice_yahoo("BRL=X", "Dólar BRL/USD")

        # Índices de mercado — mês fechado, mês atual, ano, 12M, 24M e 36M.
        # Observação: S&P 500, Dow Jones e Nasdaq são índices em pontos.
        # A conversão para BRL usa o dólar PTAX quando disponível.
        ibov_indice   = self._resumo_indice_yahoo("^BVSP", "Ibovespa")

        # ★ v17.3: IFIX — Índice de Fundos de Investimento Imobiliário
        # Ticker ^IFIX existe no Yahoo Finance; se indisponível, retorna dict vazio (graceful)
        try:
            ifix_indice = self._resumo_indice_yahoo("^IFIX", "IFIX")
            if ifix_indice.get("erro"):
                log(f"  [IFIX] Indisponível no Yahoo — campo virá vazio no JSON")
                ifix_indice = {"nome": "IFIX", "ticker": "^IFIX", "fonte": "Yahoo Finance",
                               "erro": "indisponível", "nota": "Adicionar ao robô quando ticker disponível"}
        except Exception as e_ifix:
            log(f"  [IFIX] Erro ao buscar: {e_ifix}")
            ifix_indice = {"nome": "IFIX", "ticker": "^IFIX", "fonte": "Yahoo Finance", "erro": str(e_ifix)}

        sp500_indice  = self._resumo_indice_yahoo("^GSPC", "S&P 500")
        dow_indice    = self._resumo_indice_yahoo("^DJI", "Dow Jones")
        nasdaq_indice = self._resumo_indice_yahoo("^IXIC", "Nasdaq")

        sp500_indice  = self._converter_indice_usd_para_brl(sp500_indice, dolar_indice)
        dow_indice    = self._converter_indice_usd_para_brl(dow_indice, dolar_indice)
        nasdaq_indice = self._converter_indice_usd_para_brl(nasdaq_indice, dolar_indice)

        # Focus
        focus_data = buscar_focus(self.headers)

        return {
            "atualizado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),

            "cards": {
                "selic_meta": {
                    "valor":    selic_meta,
                    "unidade":  "% a.a.",
                    "historico": selic_historico,
                },
                "cdi": {
                    "valor":   round(selic_meta - 0.10, 4) if selic_meta else None,
                    "valor_estimado": True,
                    "fonte_valor": "Referência estimada: Selic Meta - 0,10 p.p.; CDI mensal/acumulado: BCB SGS 4391",
                    "mensal":  cdi_mensal_atual,
                    "mes_ref": cdi_acum["mes_ref"] if cdi_acum else None,
                    "parcial_mes_atual": cdi_acum.get("parcial_mes_atual") if cdi_acum else None,
                    "parcial_ref":       cdi_acum.get("parcial_ref") if cdi_acum else None,
                    "acum_ano":          cdi_acum.get("acum_ano") if cdi_acum else None,
                    "acum_ano_com_parcial": cdi_acum.get("acum_ano_com_parcial") if cdi_acum else None,
                    "acum_12m": cdi_acum["acum_12m"] if cdi_acum else None,
                    "acum_24m": cdi_acum["acum_24m"] if cdi_acum else None,
                    "acum_36m": cdi_acum["acum_36m"] if cdi_acum else None,
                    "historico": cdi_acum.get("historico") if cdi_acum else [],
                    "fonte_acum": cdi_acum["fonte"] if cdi_acum else "indisponível",
                    "unidade": "% a.a.",
                },
                "ipca": {
                    "ultimo_mes":  ipca_ultimo_mes,
                    "label_mes":   ipca_label_mes,
                    "acum_ano":    ipca_acum_ano,
                    "acum_12m":    ipca_acum_12m,
                    "acum_24m":    ipca_acum_24m,
                    "acum_36m":    ipca_acum_36m,
                    "historico":   ipca_historico,
                    "meta_central": 3.0,
                    "meta_superior": 4.5,
                    "meta_inferior": 1.5,
                    "unidade": "%",
                },
                # ★ v17.1 — Poupança nova e antiga completas (mensal + acum. ano)
                "poupanca_nova": {
                    "valor":        poup_detalhado["nova"]["valor"]    if poup_detalhado else poupanca_nova,
                    "mensal":       poup_detalhado["nova"]["mensal"]   if poup_detalhado else poupanca_nova,
                    "acum_ano":     poup_detalhado["nova"]["acum_ano"] if poup_detalhado else None,
                    "historico_ano": poup_detalhado["nova"]["historico_ano"] if poup_detalhado else [],
                    "unidade":      "% a.m.",
                    "nota":         poup_detalhado["nova"]["nota"]     if poup_detalhado else "série BCB 196",
                },
                "poupanca_antiga": {
                    "valor":        poup_detalhado["antiga"]["valor"]    if poup_detalhado else None,
                    "mensal":       poup_detalhado["antiga"]["mensal"]   if poup_detalhado else None,
                    "acum_ano":     poup_detalhado["antiga"]["acum_ano"] if poup_detalhado else None,
                    "historico_ano": poup_detalhado["antiga"]["historico_ano"] if poup_detalhado else [],
                    "unidade":      "% a.m.",
                    "nota":         "TR + 0,50% a.m. — independente da Selic",
                },

                # Compatibilidade com a index atual + campos novos
                "ibovespa": {
                    "atual":    ibov_indice.get("fechamento_atual"),
                    "anterior": ibov_indice.get("fechamento_mes_anterior"),

                    # Mantém nome antigo usado pela index: usa mês atual se houver parcial; senão usa mês fechado.
                    "variacao_mensal": ibov_indice.get("variacao_mes_atual") if ibov_indice.get("variacao_mes_atual") is not None else ibov_indice.get("variacao_mes_fechado"),

                    # Novos campos
                    "fechamento_mes_anterior": ibov_indice.get("fechamento_mes_anterior"),
                    "fechamento_mes_base":     ibov_indice.get("fechamento_mes_base"),
                    "variacao_mes_fechado":    ibov_indice.get("variacao_mes_fechado"),
                    "variacao_mes_atual":      ibov_indice.get("variacao_mes_atual"),
                    "acum_ano":                ibov_indice.get("acum_ano"),
                    "acum_12m":                ibov_indice.get("acum_12m"),
                    "acum_24m":                ibov_indice.get("acum_24m"),
                    "acum_36m":                ibov_indice.get("acum_36m"),
                    "mes_anterior_label":      ibov_indice.get("mes_anterior_label"),
                    "mes_base_label":          ibov_indice.get("mes_base_label"),
                    "data_mes_anterior":       ibov_indice.get("data_mes_anterior"),
                    "data_atual":              ibov_indice.get("data_atual"),
                    "fonte":                   ibov_indice.get("fonte"),
                },

                "dolar": {
                    "atual":    dolar_indice.get("fechamento_atual"),
                    "anterior": dolar_indice.get("fechamento_mes_anterior"),

                    # Mantém nome antigo usado pela index: usa mês atual se houver parcial; senão usa mês fechado.
                    "variacao_mensal": dolar_indice.get("variacao_mes_atual") if dolar_indice.get("variacao_mes_atual") is not None else dolar_indice.get("variacao_mes_fechado"),

                    # Novos campos
                    "fechamento_mes_anterior": dolar_indice.get("fechamento_mes_anterior"),
                    "fechamento_mes_base":     dolar_indice.get("fechamento_mes_base"),
                    "variacao_mes_fechado":    dolar_indice.get("variacao_mes_fechado"),
                    "variacao_mes_atual":      dolar_indice.get("variacao_mes_atual"),
                    "acum_ano":                dolar_indice.get("acum_ano"),
                    "acum_12m":                dolar_indice.get("acum_12m"),
                    "acum_24m":                dolar_indice.get("acum_24m"),
                    "acum_36m":                dolar_indice.get("acum_36m"),
                    "mes_anterior_label":      dolar_indice.get("mes_anterior_label"),
                    "mes_base_label":          dolar_indice.get("mes_base_label"),
                    "data_mes_anterior":       dolar_indice.get("data_mes_anterior"),
                    "data_atual":              dolar_indice.get("data_atual"),
                    "fonte":                   dolar_indice.get("fonte"),
                },
            },

            # Nova estrutura principal para a tabela de mercado da página.
            "indices_mercado": {
                "dolar":     dolar_indice,
                "ibovespa":  ibov_indice,
                "ifix":      ifix_indice,   # ★ v17.3: FIIs — usado pelo fechamento rápido mobile
                "sp500":     sp500_indice,
                "dow_jones": dow_indice,
                "nasdaq":    nasdaq_indice,
            },

            # Compatibilidade com a estrutura antiga da index.
            "indices_internacionais": {
                "sp500_usd": sp500_indice.get("fechamento_atual"),
                "sp500_brl": sp500_indice.get("fechamento_atual_brl"),

                "dow_jones": dow_indice.get("fechamento_atual"),
                "nasdaq":    nasdaq_indice.get("fechamento_atual"),

                # Novos detalhamentos
                "sp500": sp500_indice,
                "dow_jones_detalhado": dow_indice,
                "nasdaq_detalhado": nasdaq_indice,
            },

            "ptax_historico": ptax_historico,
            "focus": focus_data,
            "historico_selic": selic_historico,
            "meta_vs_inflacao_efetiva": inflacao_meta_efetiva,
        }

# ---------------------------------------------------------------------------
# SIPII scraping — inalterado v15
# ---------------------------------------------------------------------------
def configurar_driver(headless=True):
    opt = webdriver.ChromeOptions()
    if headless: opt.add_argument("--headless=new")
    for arg in ["--start-maximized","--window-size=1600,1200","--no-sandbox","--disable-dev-shm-usage","--disable-gpu"]:
        opt.add_argument(arg)
    opt.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/123.0.0.0 Safari/537.36")
    return webdriver.Chrome(options=opt)

def esperar_ajax(driver, timeout=20):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return (window.jQuery ? jQuery.active === 0 : true);"))
    except: time.sleep(2)

def finalizar_driver(driver):
    try:
        if driver: driver.quit()
    except: pass

def clicar_elemento(driver, el):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.5)
    try: el.click()
    except: driver.execute_script("arguments[0].click();", el)

def clicar_por_texto(driver, texto_alvo):
    alvo = normalizar(texto_alvo)
    for a in driver.find_elements(By.TAG_NAME, "a"):
        if a.is_displayed() and normalizar(a.text.strip()) == alvo:
            clicar_elemento(driver, a); return
    raise NoSuchElementException(f"Link não encontrado: {texto_alvo}")

def abrir_site_e_preparar(driver, sigla, segmento):
    log(f"[{sigla}] Abrindo SIPII...")
    driver.get(URL_SIPII); time.sleep(3)
    clicar_por_texto(driver, segmento); esperar_ajax(driver)
    consultar = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "btn-consultar")))
    clicar_elemento(driver, consultar); esperar_ajax(driver); time.sleep(3)

def descobrir_categorias(driver):
    cats = []
    for aba in driver.find_elements(By.CSS_SELECTOR, "ul.ui-tabs-nav li a"):
        texto = aba.text.strip()
        if texto:
            norm = normalizar(texto)
            cats.append({"texto_tela": texto,
                         "csv": TEXTO_PARA_CSV.get(norm) or norm.replace(" ","_")[:31]})
    return cats

def localizar_tabela_ativa(driver):
    for p in driver.find_elements(By.CSS_SELECTOR, "div.ui-tabs-panel"):
        if p.is_displayed(): return p.find_element(By.CSS_SELECTOR, "table")
    raise Exception("Tabela ativa não encontrada")

def extrair_dados_tabela(driver, nome_csv, sigla):
    tabela = localizar_tabela_ativa(driver)
    linhas = tabela.find_elements(By.CSS_SELECTOR, "tbody tr")
    dados = []
    for i, tr in enumerate(linhas):
        tds = tr.find_elements(By.XPATH, "./td")
        if DEBUG_COLUNAS and i == 0:
            print("\n" + "="*60)
            for idx, td in enumerate(tds): print(f"  [{idx}] → '{td.text.strip()}'")
            print("="*60); return []
        if len(tds) >= 10:
            nome = tds[0].text.strip()
            if nome:
                dados.append({
                    "Categoria":        nome_csv,
                    "Fundo":            nome,
                    "Fundo_norm":       normalizar(nome),
                    "Data Inicio":      tds[1].text.strip(),
                    "Cota (R$)":        tds[3].text.strip(),
                    "Variacao Dia (%)": tds[4].text.strip(),
                    "Acum. Mes (%)":    tds[5].text.strip(),
                    "Acum. Ano (%)":    tds[6].text.strip(),
                    "Acum. 12M (%)":    tds[7].text.strip(),
                    "PL (milhoes R$)":  tds[8].text.strip(),
                    "Perfil":           sigla,
                })
    return dados

def coletar_aba(driver, sigla, segmento, cat, dados):
    try:
        clicar_por_texto(driver, cat["texto_tela"]); esperar_ajax(driver); time.sleep(1.5)
        res = extrair_dados_tabela(driver, cat["csv"], sigla)
        dados.extend(res)
        log(f"  [{sigla}] {cat['csv']}: {len(res)} fundos.")
    except (InvalidSessionIdException, WebDriverException):
        log(f"  [{sigla}] Sessão perdida em '{cat['csv']}'. Reiniciando...")
        finalizar_driver(driver)
        driver = configurar_driver(headless=True)
        try:
            abrir_site_e_preparar(driver, sigla, segmento)
            clicar_por_texto(driver, cat["texto_tela"]); esperar_ajax(driver); time.sleep(1.5)
            res = extrair_dados_tabela(driver, cat["csv"], sigla)
            dados.extend(res)
            log(f"  [{sigla}] {cat['csv']} (recuperado): {len(res)} fundos.")
        except Exception as e2:
            log(f"  [{sigla}] Falha ao recuperar '{cat['csv']}': {e2}")
    except Exception as e:
        log(f"  [{sigla}] Erro em '{cat['csv']}': {e}")
    return driver

def processar_perfil(perfil, headless=True):
    sigla, segmento = perfil["sigla"], perfil["segmento"]
    driver, dados = None, []
    try:
        driver = configurar_driver(headless=headless)
        abrir_site_e_preparar(driver, sigla, segmento)
        for cat in descobrir_categorias(driver):
            driver = coletar_aba(driver, sigla, segmento, cat, dados)
    except Exception as e:
        log(f"[{sigla}] Erro geral: {e}"); traceback.print_exc()
    finally:
        finalizar_driver(driver)
    return dados

def consolidar(todos):
    if not todos: return pd.DataFrame()
    df = pd.DataFrame(todos)
    consolidado = []
    for (fn, cat), group in df.groupby(["Fundo_norm","Categoria"], sort=False):
        reg = group.iloc[0].to_dict()
        reg["Perfis"] = " | ".join(sorted(group["Perfil"].unique()))
        consolidado.append(reg)
    return pd.DataFrame(consolidado)

def salvar_excel(df, caminho):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Fundos CAIXA"
        cols = [c for c in df.columns if c not in ("Fundo_norm","Perfil")]
        hf=PatternFill("solid",fgColor="0A1628"); hfont=Font(bold=True,color="C9A84C",size=10)
        ha=Alignment(horizontal="center",vertical="center",wrap_text=True)
        hb=Border(bottom=Side(style="medium",color="C9A84C"))
        ws.append(cols)
        for ci,cn in enumerate(cols,1):
            c=ws.cell(1,ci); c.fill=hf; c.font=hfont; c.alignment=ha; c.border=hb
        gf=Font(color="2EC27E",size=9); rf=Font(color="E05555",size=9); nf=Font(size=9)
        af=PatternFill("solid",fgColor="0F2040"); mf=PatternFill("solid",fgColor="0A1628")
        pc={"Variacao Dia (%)","Acum. Mes (%)","Acum. Ano (%)","Acum. 12M (%)"}
        for ri,row in enumerate(df[cols].itertuples(index=False),2):
            fill=af if ri%2==0 else mf
            for ci,(cn,val) in enumerate(zip(cols,row),1):
                vs=str(val) if val is not None else ""
                cell=ws.cell(ri,ci,vs); cell.fill=fill; cell.alignment=Alignment(vertical="center")
                if cn in pc and vs not in("-","—",""):
                    try:
                        n=float(vs.replace("%","").replace(" ","").replace(".","").replace(",","."))
                        cell.font=gf if n>0 else(rf if n<0 else nf)
                    except: cell.font=nf
                else: cell.font=nf
        widths={
            "Categoria":18,"Fundo":55,"Data Inicio":13,"Cota (R$)":16,
            "Variacao Dia (%)":14,"Acum. Mes (%)":13,"Acum. Ano (%)":13,
            "Acum. 12M (%)":13,"PL (milhoes R$)":18,"Perfis":30,"URL":60,
            "CNPJ":22,"codfundo":12,"Perfil de Risco":16,"Taxa Adm (%)":14,
            "Aplicacao Minima (R$)":22,"Aplicacao Adicional Minima (R$)":25,
            "Resgate Minimo (R$)":20,"Saldo Minimo (R$)":18,
            "Conversao Aplicacao":20,"Conversao Resgate":18,"Pagamento Resgate":18,
            "Benchmark Oficial":18,"Estratégia":34,"Status de Captação":22,
            "Classificação Tributária":22,"Horário Limite Aplicação":22,
            "Horário Limite Resgate":22,"Adiantamento de Resgate":28,
            "Público Alvo":34,"Segmentos":34,"Observação Operacional":70,
        }
        for ci,cn in enumerate(cols,1):
            ws.column_dimensions[get_column_letter(ci)].width=widths.get(cn,15)
        ws.row_dimensions[1].height=30; ws.freeze_panes="A2"; wb.save(caminho)
        log(f"Excel salvo: {caminho}")
    except ImportError: log("AVISO: pip install openpyxl")
    except Exception as e: log(f"Erro Excel: {e}"); traceback.print_exc()

# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------
def executar():
    log("=" * 65)
    log("⚡ ROBÔ SIPII v18.0 — atualização integral e automática do catálogo")
    log("=" * 65)

    # 1. URLs dinâmicas do portal CAIXA
    links_dinamicos = raspar_urls_caixa()

    # 2. Metadados do fundos.json (CNPJ, taxa adm, prazos, codfundo)
    #    ★ v16: também salva fundos_caixa.json para o HTML
    headers_http = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    indice_fundos_json = buscar_fundos_json(headers_http)

    # 3. Scraping SIPII por segmento
    todos_dados = []
    for perf in PERFIS:
        log(f"Raspando segmento: {perf['segmento']}...")
        todos_dados.extend(processar_perfil(perf, headless=True))
        time.sleep(2)

    # 4. Consolidação (com fallback se SIPII falhar)
    df_consolidado = consolidar(todos_dados)
    if df_consolidado.empty:
        log("[AVISO] SIPII sem dados. Ativando fallback...")
        df_consolidado = carregar_fallback_sipii()
        if df_consolidado.empty:
            log("[ERRO CRÍTICO] Fallback também vazio. Pipeline interrompido.")
            return
        log(f"[FALLBACK] Continuando com {len(df_consolidado)} fundos.")

    # 5. Enriquecimento (URL + CNPJ + codfundo + taxa adm + prazos)
    df_consolidado["URL"] = df_consolidado["Fundo"].apply(
        lambda x: encontrar_url(x, links_dinamicos))
    df_consolidado = enriquecer_dados_com_fundos_json(df_consolidado, indice_fundos_json)

    # 6. Limpeza de nomes
    df_consolidado['Fundo']      = df_consolidado['Fundo'].astype(str).apply(
        lambda x: re.sub(r'\s*\(\d+\)', '', x).strip())
    df_consolidado['Fundo_norm'] = df_consolidado['Fundo_norm'].astype(str).apply(
        lambda x: re.sub(r'\s*\(\d+\)', '', x).strip())
    for col in [
        "CNPJ", "codfundo", "Perfil de Risco", "Taxa Adm (%)",
        "Aplicacao Minima (R$)", "Aplicacao Adicional Minima (R$)",
        "Resgate Minimo (R$)", "Saldo Minimo (R$)",
        "Conversao Aplicacao", "Conversao Resgate", "Pagamento Resgate",
        "Benchmark Oficial", "Estratégia", "Status de Captação",
        "Classificação Tributária", "Classificação Investidor",
        "Horário Limite Aplicação", "Horário Limite Resgate",
        "Adiantamento de Resgate", "Modalidade Adiantamento",
        "Percentual Adiantamento (%)", "Público Alvo", "Segmentos",
        "Movimentação Automática", "Carência", "Fim Carência", "ASG",
        "Razão Social", "Observação Operacional",
        "doc_lamina", "doc_regulamento", "doc_inf_comp", "doc_comunicado",
        "doc_carta", "doc_boletim", "doc_termo", "doc_sumario", "doc_raio_x",
    ]:
        if col in df_consolidado.columns:
            df_consolidado[col] = df_consolidado[col].fillna("")

    # 7. Exportação CSV + Excel
    caminho_csv  = BASE_DIR / "dados_atuais.csv"
    caminho_xlsx = BASE_DIR / "dados_atuais.xlsx"
    df_consolidado.to_csv(caminho_csv, index=False, encoding="utf-8", quoting=csv.QUOTE_MINIMAL)
    salvar_excel(df_consolidado, caminho_xlsx)

    data_str = datetime.now().strftime("%Y%m%d")
    df_consolidado.to_csv(BASE_DIR / f"sipii_caixa_{data_str}.csv",
                          index=False, encoding="utf-8", quoting=csv.QUOTE_MINIMAL)
    salvar_excel(df_consolidado, BASE_DIR / f"sipii_caixa_{data_str}.xlsx")

    # 8. KPIs do dashboard
    gerar_json_kpis_dashboard(df_consolidado, BASE_DIR / "kpis_dashboard.json")

    # 9. Indicadores macro + Focus + CDI/IPCA 24M/36M + PTAX histórico
    try:
        coletor     = ColetorMercado()
        indicadores = coletor.coletar_todos()
        caminho_json = BASE_DIR / "mercado_atual.json"
        with open(caminho_json, "w", encoding="utf-8") as f:
            json.dump(indicadores, f, indent=4, ensure_ascii=False)

        cdi    = indicadores.get("cards", {}).get("cdi", {})
        ipca   = indicadores.get("cards", {}).get("ipca", {})
        ptax   = indicadores.get("ptax_historico", [])
        pnova  = indicadores.get("cards", {}).get("poupanca_nova", {})
        pantiga= indicadores.get("cards", {}).get("poupanca_antiga", {})
        log(f"[SUCESSO] mercado_atual.json exportado")
        log(f"  CDI: mensal={cdi.get('mensal')}% | 12M={cdi.get('acum_12m')}% | 24M={cdi.get('acum_24m')}% | 36M={cdi.get('acum_36m')}%")
        log(f"  IPCA: 12M={ipca.get('acum_12m')}% | 24M={ipca.get('acum_24m')}% | 36M={ipca.get('acum_36m')}%")
        log(f"  PTAX: {len(ptax)} fechamentos mensais pré-salvos")
        log(f"  Poupança nova:  mensal={pnova.get('valor')}% | acum. ano={pnova.get('acum_ano')}%")
        log(f"  Poupança antiga: mensal={pantiga.get('valor')}% | acum. ano={pantiga.get('acum_ano')}%")
    except Exception as e:
        log(f"[ERRO] Falha nos indicadores macro: {e}")
        traceback.print_exc()

    # 10. Limpeza de backups (mantém 5)
    limpar_backups_antigos(manter=5)

    log("=" * 65)
    log(f"[FIM] {len(df_consolidado)} fundos consolidados. Pipeline v18.0 automático completo.")
    log("=" * 65)


if __name__ == "__main__":
    executar()
